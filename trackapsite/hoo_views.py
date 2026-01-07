from django.shortcuts import render,redirect, HttpResponse, get_object_or_404
from django.urls import path, include, reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.hashers import check_password
from django.contrib import messages
from app.models import CustomUser, Event, School_Year,Announcement, Salutation,Organization, MemberType, MembershipType, Member, OfficerType, Region, Membership, Member_Event_Registration, Bulk_Event_Reg, Tags, Intetrested_Topics, IT_Topics
from django.utils.safestring import mark_safe
from django.utils import timezone
import json
from django.http import JsonResponse, Http404
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail, EmailMessage, EmailMultiAlternatives
from email.mime.image import MIMEImage
import os
import qrcode
from django.conf import settings
from django.utils.crypto import get_random_string
import datetime
import os
from django.db import transaction
from django.db.models import F, Count, Max
from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils.http import http_date

# Create your views here.
def COMPUTE_EVENT_PROGRESS(event):
    """
    Compute registration metrics for an event.

    Returns a tuple: (registered_count, available_slots, progress_percent)

    - Uses `event.max_attendees` as the maximum count and
      `event.available_slots` as the remaining slots.
    - Only computes a percentage when `event.status == 'active'` and
      `max_attendees` is present and non-zero.
    - Values are clamped to sensible integers (no negatives).
    """
    if not event:
        return 0, 0, 0

    # Only compute progress for active events
    if getattr(event, 'status', None) != 'active':
        avail = getattr(event, 'available_slots', 0) or 0
        return 0, int(avail), 0

    max_att = getattr(event, 'max_attendees', None)
    avail = getattr(event, 'available_slots', None)

    # If max_attendees is not provided or zero, can't compute percent
    try:
        if max_att is None or int(max_att) == 0:
            return 0, int(avail) if avail is not None else 0, 0
    except Exception:
        return 0, int(avail) if avail is not None else 0, 0

    # If available_slots missing, assume 0 (fully booked) to avoid None math
    try:
        avail_i = int(avail) if avail is not None else 0
    except Exception:
        avail_i = 0

    try:
        max_i = int(max_att)
    except Exception:
        return 0, avail_i, 0

    registered = max(0, max_i - avail_i)
    progress_percent = int(round((registered / max_i) * 100)) if max_i > 0 else 0
    return registered, avail_i, progress_percent
@login_required(login_url='/')
def home(request):
    # Get all members and events
    members = Member.objects.all()
    events = Event.objects.all()

    # Fetch the latest active event for the progress bar
    event = Event.objects.filter(status='active').order_by('-date').first()

    # Compute progress metrics using helper (based on max_attendees and available_slots)
    if event:
        registered_count, available_slots, progress_percent = COMPUTE_EVENT_PROGRESS(event)
    else:
        registered_count, available_slots, progress_percent = 0, 0, 0

    # expose available school years for client-side filtering
    school_years = School_Year.objects.all().order_by('-sy_start')

    context = {
        'members': members,
        'events': events,
        'event': event,
        'registered_count': registered_count,
        'available_slots': available_slots,
        'progress_percent': progress_percent,  # pass percentage
        'school_years': school_years,
    }

    # Serialize events for the client-side calendar (ISO date strings)
    events_json = []
    for ev in events:
        # Serialize to date-only string (YYYY-MM-DD) to avoid timezone shifts
        date_iso = None
        if getattr(ev, 'date', None):
            try:
                d = ev.date
                # If it's a datetime, convert to date first to avoid timezone conversion in the browser
                if hasattr(d, 'date'):
                    date_iso = d.date().isoformat()
                else:
                    date_iso = d.isoformat()
            except Exception:
                date_iso = str(ev.date)

        # compute available slots for each event
        # Do not calculate registration totals here; expose stored available_slots if present
        try:
            available_slots = getattr(ev, 'available_slots', None)
        except Exception:
            available_slots = None

        # banner url (if ImageField)
        banner_url = None
        try:
            b = getattr(ev, 'banner', None)
            if b:
                # ImageField has .url
                banner_url = getattr(b, 'url', str(b))
        except Exception:
            banner_url = None

        # chair/co-chair names
        chair_name = None
        co_chair_name = None
        try:
            if getattr(ev, 'chair_id', None):
                cu = CustomUser.objects.filter(id=ev.chair_id).first()
                if cu:
                    chair_name = f"{getattr(cu, 'first_name', '')} {getattr(cu, 'last_name', '')}".strip()
            if getattr(ev, 'co_chair_id', None):
                cu2 = CustomUser.objects.filter(id=ev.co_chair_id).first()
                if cu2:
                    co_chair_name = f"{getattr(cu2, 'first_name', '')} {getattr(cu2, 'last_name', '')}".strip()
        except Exception:
            chair_name = chair_name or None
            co_chair_name = co_chair_name or None

        events_json.append({
            'id': ev.id,
            'title': getattr(ev, 'title', '') or '',
            'theme': getattr(ev, 'theme', '') or '',
            'date': date_iso,
            'banner_url': banner_url,
            'chair_id': getattr(ev, 'chair_id', None),
            'chair_name': chair_name,
            'co_chair_id': getattr(ev, 'co_chair_id', None),
            'co_chair_name': co_chair_name,
            'max_attendees': getattr(ev, 'max_attendees', None),
            'available_slots': available_slots,
            'status': getattr(ev, 'status', None),
        })

    context['events_json'] = events_json

    # ----- Member rankings by consecutive attendance -----
    try:
        # Allow optional query param `school_year` to filter rankings by school year
        selected_sy = request.GET.get('school_year')

        ev_qs = Event.objects.filter(date__isnull=False)
        if selected_sy:
            try:
                ev_qs = ev_qs.filter(school_year_id=int(selected_sy))
            except Exception:
                pass

        # Order events by date (ignore events without a date)
        ordered_event_ids = list(ev_qs.order_by('date').values_list('id', flat=True))

        # Map member_id -> set(event_id) for approved member registrations
        regs = Member_Event_Registration.objects.filter(event_id__in=ordered_event_ids, is_approved=True).values('member_id_id', 'event_id')
        member_events_map = {}
        for r in regs:
            mid = r.get('member_id_id')
            eid = r.get('event_id')
            if mid is None or eid is None:
                continue
            member_events_map.setdefault(mid, set()).add(eid)

        member_rankings = []
        for m in members:
            mid = getattr(m, 'id', None)
            events_set = member_events_map.get(mid, set())

            # Compute longest consecutive streak across ordered events
            longest = 0
            curr = 0
            for eid in ordered_event_ids:
                if eid in events_set:
                    curr += 1
                    if curr > longest:
                        longest = curr
                else:
                    curr = 0

            total_joined = len(events_set)

            # Friendly member name and organization if available
            try:
                first = getattr(m, 'first_name', '') or ''
                last = getattr(m, 'last_name', '') or ''
                name = (first + ' ' + last).strip() or str(m)
            except Exception:
                name = str(m)
            org_name = ''
            try:
                org = getattr(m, 'organization', None)
                if org:
                    org_name = getattr(org, 'name', '') or str(org)
            except Exception:
                org_name = ''

            member_rankings.append({
                'member': m,
                'member_name': name,
                'organization': org_name,
                'longest_streak': longest,
                'total_joined': total_joined,
            })

        # sort by longest streak desc, then by total joined desc
        member_rankings.sort(key=lambda x: (x['longest_streak'], x['total_joined']), reverse=True)

        # Counts for thresholds (3x, 5x, 10x in a row) - use friendly keys for template
        streak_counts = {
            'three': sum(1 for r in member_rankings if r['longest_streak'] >= 3),
            'five': sum(1 for r in member_rankings if r['longest_streak'] >= 5),
            'ten': sum(1 for r in member_rankings if r['longest_streak'] >= 10),
        }
    except Exception:
        member_rankings = []
        streak_counts = {'three': 0, 'five': 0, 'ten': 0}

    context['member_rankings'] = member_rankings
    context['streak_counts'] = streak_counts
    context['selected_school_year'] = int(selected_sy) if selected_sy and str(selected_sy).isdigit() else None

    # Build column chart data: categories = event titles; series = Registered / Attended counts
    categories = []
    registered_counts = []
    attended_counts = []
    for ev in events:
        try:
            ev_id = getattr(ev, 'id', None)
            title = getattr(ev, 'title', '') or f"Event {ev_id}"
            # Member registrations
            mem_registered = Member_Event_Registration.objects.filter(event_id=ev_id, is_approved=True).count()
            mem_attended = Member_Event_Registration.objects.filter(event_id=ev_id, is_present=True).count()
            # Bulk registrations
            bulk_registered = Bulk_Event_Reg.objects.filter(event_id=ev_id, is_approved=True).count()
            bulk_attended = Bulk_Event_Reg.objects.filter(event_id=ev_id, is_present=True).count()

            total_registered = mem_registered + bulk_registered
            total_attended = mem_attended + bulk_attended

        except Exception:
            total_registered = 0
            total_attended = 0
            title = getattr(ev, 'title', '') or ''

        categories.append(title)
        registered_counts.append(total_registered)
        attended_counts.append(total_attended)

    # Series structure expected by ApexCharts
    chart_series = [
        {'name': 'Registered', 'data': registered_counts},
        {'name': 'Attended', 'data': attended_counts},
    ]

    context['chart_series_json'] = json.dumps(chart_series)
    context['chart_categories_json'] = json.dumps(categories)

    # Build pie chart data for Interested IT Topics: labels = IT_Topics.name, series = counts
    try:
        topic_qs = Intetrested_Topics.objects.values('topic_id').annotate(cnt=Count('id')).order_by('-cnt')
        pie_labels = []
        pie_series = []
        for row in topic_qs:
            tid = row.get('topic_id')
            cnt = row.get('cnt') or 0
            name = IT_Topics.objects.filter(id=tid).values_list('name', flat=True).first() or f"Topic {tid}"
            pie_labels.append(name)
            pie_series.append(cnt)
    except Exception:
        pie_labels = []
        pie_series = []

    context['topics_pie_series_json'] = json.dumps(pie_series)
    context['topics_pie_labels_json'] = json.dumps(pie_labels)

    return render(request, 'hoo/home.html', context)


@login_required(login_url='/')
def GENERATE_REPORT(request):
    """Generate Excel report of members for selected school year range.

    Expects POST with `sy_start` and `sy_end` (School_Year ids).
    Builds an .xlsx with columns:
      A: No. (serial starting at 1)
      B: School Name (member.organization.name)
      C: Member Email (member.admin.email)
      D: Member Lastname (member.admin.last_name)
      E: Member Firstname (member.admin.first_name)
    """
    if request.method != 'POST':
        return HttpResponse(status=405)

    sy_start_id = request.POST.get('sy_start')
    sy_end_id = request.POST.get('sy_end')

    # Determine selected school years; fall back to all if none provided
    sy_qs = School_Year.objects.all()
    try:
        if sy_start_id and sy_end_id:
            sy1 = School_Year.objects.filter(id=int(sy_start_id)).first()
            sy2 = School_Year.objects.filter(id=int(sy_end_id)).first()
            if sy1 and sy2:
                start_date = min(sy1.sy_start, sy2.sy_start)
                end_date = max(sy1.sy_end, sy2.sy_end)
                sy_qs = School_Year.objects.filter(sy_start__gte=start_date, sy_end__lte=end_date)
            elif sy1:
                sy_qs = School_Year.objects.filter(id=sy1.id)
            elif sy2:
                sy_qs = School_Year.objects.filter(id=sy2.id)
    except Exception:
        sy_qs = School_Year.objects.all()

    # Get members who have a Membership in the selected school years
    memberships = Membership.objects.filter(school_year__in=sy_qs)
    member_ids = list(memberships.values_list('member_id', flat=True)) if memberships.exists() else []
    members = Member.objects.filter(id__in=member_ids).distinct() if member_ids else Member.objects.none()

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members'

    # Header row
    ws['A1'] = 'No.'
    ws['B1'] = 'School Name'
    ws['C1'] = 'Member Email'
    ws['D1'] = 'Member Lastname'
    ws['E1'] = 'Member Firstname'

    # Bold requested header cells (School Name, Member Email, Lastname, Firstname)
    bold_font = Font(bold=True)
    ws['B1'].font = bold_font
    ws['C1'].font = bold_font
    ws['D1'].font = bold_font
    ws['E1'].font = bold_font

    # Events as additional columns starting at F (column 6).
    events_qs = Event.objects.filter(school_year__in=sy_qs).order_by('date')
    event_list = list(events_qs)
    start_col = 6  # 'F'
    for i, ev in enumerate(event_list):
        cell = ws.cell(row=1, column=start_col + i, value=(getattr(ev, 'title', '') or ''))
        cell.font = bold_font

    # Preload member-event registrations for quick lookup
    reg_map = {}
    if event_list and member_ids:
        ev_ids = [e.id for e in event_list]
        regs = Member_Event_Registration.objects.filter(event_id__in=ev_ids, member_id_id__in=member_ids).values('event_id', 'member_id_id', 'is_present')
        for r in regs:
            reg_map[(r.get('member_id_id'), r.get('event_id'))] = r.get('is_present')

    # Fill rows
    row = 2
    for idx, m in enumerate(members, start=1):
        # Resolve school name from organization_id to avoid relation surprises
        school_name = ''
        try:
            org_id = getattr(m, 'organization_id', None) or (getattr(m, 'organization', None) and getattr(m.organization, 'id', None))
            if org_id:
                school_name = Organization.objects.filter(id=org_id).values_list('name', flat=True).first() or ''
        except Exception:
            school_name = ''

        # Resolve member info from admin_id (CustomUser)
        email = ''
        first = ''
        last = ''
        try:
            admin_id = getattr(m, 'admin_id', None) or (getattr(m, 'admin', None) and getattr(m.admin, 'id', None))
            if admin_id:
                cu = CustomUser.objects.filter(id=admin_id).values('email', 'first_name', 'last_name').first()
                if cu:
                    email = cu.get('email', '') or ''
                    first = cu.get('first_name', '') or ''
                    last = cu.get('last_name', '') or ''
        except Exception:
            pass

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=school_name)
        ws.cell(row=row, column=3, value=email)
        ws.cell(row=row, column=4, value=last)
        ws.cell(row=row, column=5, value=first)
        # Fill attendance columns per event: 'Attended' if is_present truthy, otherwise 'Not Attended'
        for j, ev in enumerate(event_list):
            present = reg_map.get((m.id, ev.id))
            attended_text = 'Attended' if bool(present) else 'Not Attended'
            ws.cell(row=row, column=start_col + j, value=attended_text)
        row += 1

    # Prepare response
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"members_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='/')
def EVENT_ANALYTICS(request):
    """Simple placeholder view for Event Analytics page."""
    # Pass the latest active event to the template for display
    event = Event.objects.filter(status='active').order_by('-date').first()

    # Provide school years and events JSON for client-side filtering
    school_years = School_Year.objects.all().order_by('-sy_start')
    events = Event.objects.all().order_by('-date')

    events_json = []
    for ev in events:
        try:
            sy_id = getattr(ev, 'school_year_id', None) or getattr(ev, 'school_year', None) and getattr(ev.school_year, 'id', None)
        except Exception:
            sy_id = None
        events_json.append({
            'id': ev.id,
            'title': getattr(ev, 'title', '') or '',
            'date': ev.date.isoformat() if getattr(ev, 'date', None) and hasattr(ev.date, 'isoformat') else (str(ev.date) if getattr(ev, 'date', None) else ''),
            'status': getattr(ev, 'status', '') or '',
            'school_year_id': sy_id,
        })

    # Compute registered / attended counts for the active event (sum of member + bulk approved)
    registered_total = 0
    attended_total = 0
    # If there's an active event, compute the initial totals so the page shows accurate counts
    try:
        if event:
            mem_registered = Member_Event_Registration.objects.filter(event=event, is_approved=True).count()
            bulk_registered = Bulk_Event_Reg.objects.filter(event=event, is_approved=True).count()
            registered_total = (mem_registered or 0) + (bulk_registered or 0)

            mem_attended = Member_Event_Registration.objects.filter(event=event, is_present=True).count()
            bulk_attended = Bulk_Event_Reg.objects.filter(event=event, is_present=True).count()
            attended_total = (mem_attended or 0) + (bulk_attended or 0)
    except Exception:
        registered_total = registered_total or 0
        attended_total = attended_total or 0

    # Build organizations JSON for client-side charting (categories = initials)
    # Count registrations per organization for the selected/active event
    organizations = Organization.objects.all()
    orgs_json = []
    for org in organizations:
        try:
            if event:
                # Count Member_Event_Registration entries for this event where the
                # related Member belongs to the organization.
                member_count = Member_Event_Registration.objects.filter(
                    event=event,
                    member_id__organization_id=org.id,
                    is_approved=True,
                ).count()

                # Also include Bulk_Event_Reg entries for this event where the
                # registered_by (Member) belongs to the organization.
                try:
                    bulk_count = Bulk_Event_Reg.objects.filter(
                        event=event,
                        registered_by__organization_id=org.id,
                        is_approved=True,
                    ).count()
                except Exception:
                    bulk_count = 0

                total_count = member_count + (bulk_count or 0)
            else:
                total_count = 0
        except Exception:
            total_count = 0
        orgs_json.append({
            'id': org.id,
            'initials': getattr(org, 'initials', '') or '',
            'name': getattr(org, 'name', '') or '',
            'value': total_count,
        })

    context = {
        'event': event,
        'school_years': school_years,
        'events_json': json.dumps(events_json),
        'registered_total': registered_total,
        'attended_total': attended_total,
        'organizations_json': json.dumps(orgs_json),
    }
    return render(request, 'hoo/event_analytics.html', context)

@login_required(login_url='/')
def GET_EVENT_STATS(request, id):
    """Return JSON with registered and attended totals (member + bulk) for event `id`."""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.method == 'GET':
        try:
            ev_id = int(id)
        except Exception:
            return JsonResponse({'error': 'invalid id'}, status=400)

        try:
            mem_registered = Member_Event_Registration.objects.filter(event_id=ev_id, is_approved=True).count()
            bulk_registered = Bulk_Event_Reg.objects.filter(event_id=ev_id, is_approved=True).count()
            registered_total = mem_registered + bulk_registered

            mem_attended = Member_Event_Registration.objects.filter(event_id=ev_id, is_present=True).count()
            bulk_attended = Bulk_Event_Reg.objects.filter(event_id=ev_id, is_present=True).count()
            attended_total = mem_attended + bulk_attended

            # Build per-organization counts for this event
            orgs = []
            try:
                organizations = Organization.objects.all()
                for org in organizations:
                    try:
                        mem_count = Member_Event_Registration.objects.filter(
                            event_id=ev_id,
                            member_id__organization_id=org.id,
                            is_approved=True,
                        ).count()
                    except Exception:
                        mem_count = 0
                    try:
                        bulk_count = Bulk_Event_Reg.objects.filter(
                            event_id=ev_id,
                            registered_by__organization_id=org.id,
                            is_approved=True,
                        ).count()
                    except Exception:
                        bulk_count = 0
                    orgs.append({
                        'id': org.id,
                        'initials': getattr(org, 'initials', '') or '',
                        'name': getattr(org, 'name', '') or '',
                        'value': (mem_count or 0) + (bulk_count or 0),
                    })
            except Exception:
                orgs = []

            return JsonResponse({'registered_total': registered_total, 'attended_total': attended_total, 'organizations': orgs})
        except Exception:
            return JsonResponse({'registered_total': 0, 'attended_total': 0})

    raise Http404('Invalid request')


@login_required(login_url='/')
def GET_EVENT_ATTENDING_PIE(request, id):
    """Return JSON for pie chart grouped by attendee role categories.

    Labels: 'Student', 'Professor/Teacher/Instructor', 'Other'
    Series: count of distinct `bulk_reg` ids that have an Attending in that category
    for the given event id.
    """
    if not (request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.method == 'GET'):
        raise Http404('Invalid request')

    try:
        ev_id = int(id)
    except Exception:
        return JsonResponse({'error': 'invalid id'}, status=400)

    # Prepare sets of Bulk_Event_Reg ids per category using Bulk_Event_Reg.attending_as
    student_set = set()
    prof_set = set()
    other_set = set()

    # Read directly from Bulk_Event_Reg so we use the saved attending_as string
    # Only consider approved bulk registrations
    bulk_qs = Bulk_Event_Reg.objects.filter(event_id=ev_id, is_approved=True)
    for b in bulk_qs:
        roles_str = (getattr(b, 'attending_as', '') or '').strip()
        if not roles_str:
            continue

        # support comma-separated roles
        parts = [p.strip().lower() for p in roles_str.split(',') if p.strip()]
        for role in parts:
            if 'student' in role:
                student_set.add(b.id)
            elif any(k in role for k in ('professor', 'teacher', 'instructor')):
                prof_set.add(b.id)
            else:
                other_set.add(b.id)

    labels = []
    series = []

    if student_set:
        labels.append('Student')
        series.append(len(student_set))
    if prof_set:
        labels.append('Professor/Teacher/Instructor')
        series.append(len(prof_set))
    if other_set:
        labels.append('Other')
        series.append(len(other_set))

    return JsonResponse({'labels': labels, 'series': series})


@login_required(login_url='/')
def GET_EVENT_COMPETITOR_COUNTS(request, id):
    """Return counts of competitor types for an event.

    Response: { labels: [...], series: [...] }
    Groups `Bulk_Event_Reg.if_competitor` and counts rows where `is_competitor=1`.
    """
    if not (request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.method == 'GET'):
        raise Http404('Invalid request')

    try:
        ev_id = int(id)
    except Exception:
        return JsonResponse({'error': 'invalid id'}, status=400)

    try:
        qs = Bulk_Event_Reg.objects.filter(event_id=ev_id, is_competitor=1, is_approved=True)
        grouped = qs.values('if_competitor').annotate(count=Count('id')).order_by('-count')
        labels = []
        series = []
        for item in grouped:
            val = item.get('if_competitor') or ''
            if not val.strip():
                continue
            labels.append(val)
            series.append(item.get('count', 0))
        return JsonResponse({'labels': labels, 'series': series})
    except Exception:
        return JsonResponse({'labels': [], 'series': []})



#for the profile
@login_required(login_url='/')
def PROFILE(request):
    user = CustomUser.objects.get(id = request.user.id)
    
    context = {
        "user":user,

    }
    return render(request, 'hoo/profile.html', context)



@login_required(login_url='/')
def PROFILE_UPDATE(request):
    if request.method == "POST":
        profile_pic = request.FILES.get('profile_pic')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        change_password = request.POST.get('change_password')  # renamed field

        try:
            customuser = CustomUser.objects.get(id=request.user.id)

            customuser.first_name = first_name
            customuser.last_name = last_name
            customuser.email = email
            customuser.username = username

            # If change password field is filled, update password
            if change_password and change_password.strip() != "":
                customuser.set_password(change_password)
                customuser.save()
                update_session_auth_hash(request, customuser)  # keep user logged in
            else:
                customuser.save()

            if profile_pic and profile_pic != "":
                customuser.profile_pic = profile_pic
                customuser.save()

            messages.success(request, 'Your profile was updated successfully!')
            return redirect('profile_update_hoo')  # redirect to the same update page
        except Exception as e:
            print("Error updating profile:", e)  # for debugging
            messages.error(request, 'Failed to update your profile')
            return redirect('profile_update_hoo')
    
    return render(request, 'hoo/profile.html')  # use your correct template




def PROFILE_PASSWORD_PAGE(request):
    user = CustomUser.objects.get(id = request.user.id)
    
    context = {
        "user":user,

    }
    return render(request, 'hoo/change_password.html', context)



@login_required(login_url='/')
def CHANGE_PASSWORD(request):
    if request.method == "POST":
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        repeat_new_password = request.POST.get('repeat_new_password')
        user = CustomUser.objects.get(id=request.user.id)

        # Check current password
        if not check_password(current_password, user.password):
            messages.error(request, "Current password is incorrect.")
            return redirect('change_password')

        # Check new password requirements
        errors = []
        if len(new_password) < 8:
            errors.append("Password must be at least 8 characters long.")
        if not any(c.isupper() for c in new_password):
            errors.append("Password must contain at least one capital letter.")
        if not any(c.isdigit() for c in new_password):
            errors.append("Password must contain at least one number.")
        if not any(c in '!@#$%^&*(),.?":{}|<>' for c in new_password):
            errors.append("Password must contain at least one special character.")
        if new_password != repeat_new_password:
            errors.append("Passwords do not match.")

        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect('change_password')

        # Set new password
        user.set_password(new_password)
        user.save()
        messages.success(request, "Your password has been changed successfully.")
        return redirect('login')

    return render(request, 'change_password.html')


# For Schoolyear 
def ADD_SCHOOLYEAR(request):
    if request.method == "POST":
        sy_start = request.POST.get('sy_start')
        sy_end = request.POST.get('sy_end')

        # Deactivate all existing school years
        School_Year.objects.update(status=0)

        # Create the new one as active
        school_year = School_Year(
            sy_start=sy_start,
            sy_end=sy_end,
            status=1,  # auto set active
        )
        school_year.save()

        messages.success(request, 'Cycle successfully added and set as Active!')
        return redirect('add_schoolyear')

    return render(request, 'hoo/add_schoolyear.html')

def VIEW_SCHOOLYEAR(request):
    schoolyear = School_Year.objects.all()
    
    context = {
        'schoolyear':schoolyear,
    }
    # print(teacher)
    return render(request, 'hoo/view_schoolyear.html', context)


def EDIT_SCHOOLYEAR(request, id):
    schoolyear = School_Year.objects.filter(id = id)
    
    context = {
        'schoolyear':schoolyear
    }
    
    return render(request,'hoo/edit_schoolyear.html', context)


def UPDATE_SCHOOLYEAR(request):
    if request.method == "POST":
        id = request.POST.get('id')
        sy_start = request.POST.get('sy_start')
        sy_end = request.POST.get('sy_end')
        # print(program)
        
        schoolyear = School_Year.objects.get(id = id)
        schoolyear.sy_start = sy_start
        schoolyear.sy_end = sy_end
        
        schoolyear.save()
        
        messages.success(request, "Cycle successfully updated")
        return redirect('view_schoolyear')
    return render(request, 'hoo/edit_schoolyear.html')

def DELETE_SCHOOLYEAR(request, id):
    schoolyear = School_Year.objects.get(id = id)
    schoolyear.delete()
    messages.success(request, 'Cycle successfully deleted')
    return redirect('view_schoolyear')

# For Member List
def ADD_MEMBER(request):
    salutations = Salutation.objects.all()
    membershiptypes = MembershipType.objects.all()
    # membertypes = MemberType.objects.all()
    officertypes = OfficerType.objects.all()
    organizations = Organization.objects.all()
    # customuser = CustomUser.objects.all()
    
    if request.method == "POST":
        # is_superuser = request.POST.get('is_superuser')
        # is_active = request.POST.get('is_active')
        # user_type = request.POST.get('user_type')
        
        membershiptype_id = request.POST.get('membershiptype_id')
        # membertype_id = request.POST.get('membertype_id')
        organization_id = request.POST.get('organization_id')
        
        salutation_id = request.POST.get('salutation_id')
        officertype_id = request.POST.get('officertype_id')
        first_name = request.POST.get('first_name').upper()
        last_name = request.POST.get('last_name').upper()
        middle_name = request.POST.get('middle_name').upper()
        # profile_pic = request.FILES.get('profile_pic')
        position = request.POST.get('position').upper()
        email = request.POST.get('email')
        contact_no = request.POST.get('contact_no')
        birthdate = request.POST.get('birthdate')
        
        # Get the birthdate and convert it to a date object
        birthdate_str = request.POST.get('birthdate')
        birthdate = None
        if birthdate_str:
            try:
                birthdate = datetime.datetime.strptime(birthdate_str, '%Y-%m-%d').date()
            except ValueError:
                messages.warning(request, 'Invalid birthdate format. Please use YYYY-MM-DD.')
                return redirect('add_member')
        
        facebook_profile_link = request.POST.get('facebook_profile_link')
        # proof_of_payment = request.FILES.get('proof_of_payment')
        payment_date = request.POST.get('payment_date')
        terms_accepted = request.POST.get('terms_accepted') == 'true'
        
        username = request.POST.get('username')
        password = request.POST.get('password')
                 
        print(f"Checking email: {email}, Exists: {CustomUser .objects.filter(email=email).exists()}")              
        if email and CustomUser.objects.filter(email=email).exists():
             messages.warning(request,'Email is already taken')
             return redirect('add_member')
         
        elif CustomUser.objects.filter(username=username).exists():
             messages.warning(request,'Username is already taken')
             return redirect('add_member')     
         
        else:
            user = CustomUser(
                # is_superuser = is_superuser,
                # is_active = is_active,
                first_name = first_name,
                last_name = last_name,
                username = username,
                email = email,
                user_type = 3,
                # profile_pic = profile_pic,
                )      
              
            user.set_password(password)
            user.save()
            
            member = Member(
                admin = user, 
                membershiptype_id = membershiptype_id,
                # membertype_id = membertype_id,
                officertype_id = officertype_id,
                organization_id = organization_id,
                
                salutation_id = salutation_id,
                middle_name = middle_name,
                birthdate = birthdate,
                position = position,
                contact_no = contact_no,
                facebook_profile_link = facebook_profile_link,
                payment_date = payment_date,
                # proof_of_payment = proof_of_payment,
                terms_accepted = terms_accepted,
            )
            
            # member.set_password(password)
            member.save()
            
            messages.success(request,user.first_name + " " + user.last_name + " is successfully added")
            return redirect('add_member')
    
    context = {
        'salutations': salutations,  # Pass the salutations to the template
        'membershiptypes':membershiptypes,
        # 'membertypes': membertypes,
        'officertypes': officertypes,
        'organizations': organizations,
    }
    
    return render(request, 'hoo/add_member.html', context)

def VIEWALL_MEMBER(request):
    # customuser = request.user
    customuser = CustomUser.objects.all()
    members = Member.objects.all()
    salutations = Salutation.objects.all()
    membershiptypes = MembershipType.objects.all()
    # membertypes = MemberType.objects.all()
    officertypes = OfficerType.objects.all()
    organizations = Organization.objects.all()
    
    context = {
        'customuser':customuser,
        'members':members,
        'salutations': salutations,  # Pass the salutations to the template
        'membershiptypes':membershiptypes,
        # 'membertypes': membertypes,
        'officertypes': officertypes,
        'organizations': organizations,
    }
    # print(customuser)
    return render(request, 'hoo/viewall_member.html', context)

def EDIT_MEMBER(request, id):
    salutations = Salutation.objects.all()
    membershiptypes = MembershipType.objects.all()
    # membertypes = MemberType.objects.all()
    officertypes = OfficerType.objects.all()
    organizations = Organization.objects.all()
    
    member = get_object_or_404(Member, id=id)

    # selected_user =   # Assuming 'admin' is a ForeignKey to CustomUser 
    
    context = {
        # 'selected_user': selected_user,  # Pass the selected user to the template
        'member': member,
        'salutations': salutations,  # Pass the salutations to the template
        'membershiptypes': membershiptypes,
        # 'membertypes': membertypes,
        'officertypes': officertypes,
        'organizations': organizations,
    }
    
    return render(request, 'hoo/edit_member.html', context)

def UPDATE_MEMBER(request):
    if request.method == "POST":
        member_id = request.POST.get('member_id')
        is_superuser = request.POST.get('is_superuser') == 'on'  # Convert to boolean
        is_active = request.POST.get('is_active')      # Convert to boolean
        user_type = request.POST.get('user_type')
        
        membership_type = request.POST.get('membership_type')
        # member_type = request.POST.get('member_type')
        
        salutation = request.POST.get('salutation')
        first_name = request.POST.get('first_name').upper()
        last_name = request.POST.get('last_name').upper()
        middle_name = request.POST.get('middle_name').upper()
        profile_pic = request.FILES.get('profile_pic')
        position = request.POST.get('position').upper()
        email = request.POST.get('email')
        contact_no = request.POST.get('contact_no')
        birthdate = request.POST.get('birthdate')
        facebook_profile_link = request.POST.get('facebook_profile_link')
        # proof_of_payment = request.FILES.get('proof_of_payment')
        payment_date = request.POST.get('payment_date')
        terms_accepted = request.POST.get('terms_accepted')
        
        username = request.POST.get('username')
        password = request.POST.get('password')
        # print(program)
        
        customuser = CustomUser.objects.get(id = member_id)
        # customuser = get_object_or_404(CustomUser , id=member_id)
        
        
        # customuser.id = member_id
        customuser.is_superuser = is_superuser
        customuser.is_active = is_active
        customuser.user_type = user_type
        
        customuser.membership_type = membership_type
        # customuser.member_type = 1
        
        customuser.salutation = salutation
        customuser.first_name = first_name
        customuser.last_name = last_name
        customuser.middle_name = middle_name
        customuser.profile_pic = profile_pic
        customuser.position = position
        customuser.email = email
        customuser.contact_no = contact_no
        customuser.birthdate = birthdate
        customuser.facebook_profile_link = facebook_profile_link
        customuser.payment_date = payment_date
        customuser.terms_accepted = terms_accepted
        
        customuser.username = username
        customuser.password = password
        
        
        # Retrieve the user object or return a 404 error if not found
        
        if password !=None and password != "":
            customuser.set_password(password)
            
        customuser.save()
        
        messages.success(request, "Member successfully updated")
        return redirect('viewall_member')
    return render(request, 'hoo/edit_member.html')

def DELETE_MEMBER(request, id):
    customuser = CustomUser.objects.get(id = id)
    customuser.delete()
    messages.success(request, 'Member successfully deleted')
    return redirect('viewall_member')

def MEMBER_DETAILS(request,id):
    selected_member = Member.objects.filter(id = id)
    member= Member.objects.get(id=id)
    
    context = {
        'member':member,
        'selected_member':selected_member,
   
    }
    return render(request, 'hoo/member_details.html', context)

#For Region/Chapter Modules
def VIEW_REGION(request):
    region = Region.objects.all()
    
    context = {
        'region':region,
    }
    # print(teacher)
    return render(request, 'hoo/view_region.html', context)

def ADD_REGION(request):
    if request.method == "POST":
        region_name = request.POST.get('region_name')
        region_info = request.POST.get('region_info')
        # print(program_name)
        
        region = Region (
            name = region_name,
            info = region_info,
            created_by_id=request.user.id  # Set the created_by_username to the current user
        )
        region.save()
        messages.success(request, 'Region/Chapter successfully added!')
        return redirect('add_region')
    return render(request, 'hoo/add_region.html')

def EDIT_REGION(request, id):
    region = Region.objects.filter(id = id)
    
    context = {
        'region':region,
    }
    
    return render(request,'hoo/edit_region.html', context)

def UPDATE_REGION(request):
    if request.method == "POST":
        id = request.POST.get('region_id')
        region_name = request.POST.get('region_name')
        region_info = request.POST.get('region_info')
        # print(program)
        
        region = Region.objects.get(id = id)
        region.name = region_name
        region.info = region_info
        
        region.save()
        
        messages.success(request, "Region/Chapter successfully updated")
        return redirect('view_region')
    return render(request, 'hoo/edit_region.html')

def DELETE_REGION(request, id):
    region = Region.objects.get(id = id)
    region.delete()
    messages.success(request, 'Region/Chapter successfully deleted')
    return redirect('view_region')

#For Officertypes Modules
def VIEW_OFFICERTYPE(request):
    officertype = OfficerType.objects.all()
    
    context = {
        'officertype':officertype,
    }
    # print(teacher)
    return render(request, 'hoo/view_officertype.html', context)

def ADD_OFFICERTYPE(request):
    if request.method == "POST":
        officertype_name = request.POST.get('officertype_name')
        officertype_info = request.POST.get('officertype_info')
        # print(program_name)
        
        officertype = OfficerType (
            name = officertype_name,
            info = officertype_info,
            created_by_id=request.user.id  # Set the created_by_username to the current user
        )
        officertype.save()
        messages.success(request, 'OfficerType successfully added!')
        return redirect('add_officertype')
    return render(request, 'hoo/add_officertype.html')

def DELETE_OFFICERTYPE(request, id):
    officertype = OfficerType.objects.get(id = id)
    officertype.delete()
    messages.success(request, 'OfficerType successfully deleted')
    return redirect('view_officertype')

def EDIT_OFFICERTYPE(request, id):
    officertype = OfficerType.objects.filter(id = id)
    
    context = {
        'officertype':officertype,
    }
    
    return render(request,'hoo/edit_officertype.html', context)

def UPDATE_OFFICERTYPE(request):
    if request.method == "POST":
        id = request.POST.get('officertype_id')
        officertype_name = request.POST.get('officertype_name')
        officertype_info = request.POST.get('officertype_info')
        # print(program)
        
        officertype = OfficerType.objects.get(id = id)
        officertype.name = officertype_name
        officertype.info = officertype_info
        
        officertype.save()
        
        messages.success(request, "OfficerType successfully updated")
        return redirect('view_officertype')
    return render(request, 'hoo/edit_officertype.html')

#For Membership Modules
def VIEW_MEMBERSHIPTYPE(request):
    membershiptype = MembershipType.objects.all()
    
    context = {
        'membershiptype':membershiptype,
    }
    # print(teacher)
    return render(request, 'hoo/view_membershiptype.html', context)

def ADD_MEMBERSHIPTYPE(request):
    if request.method == "POST":
        membershiptype_name = request.POST.get('membershiptype_name')
        membershiptype_info = request.POST.get('membershiptype_info')
        membershiptype_price = request.POST.get('membershiptype_price')
        # print(program_name)
        
        membershiptype = MembershipType (
            name = membershiptype_name,
            info = membershiptype_info,
            price = membershiptype_price,
            created_by_id=request.user.id  # Set the created_by_username to the current user
        )
        membershiptype.save()
        messages.success(request, 'MembershipType successfully added!')
        return redirect('view_membershiptype')
    return render(request, 'hoo/add_membershiptype.html')

def DELETE_MEMBERSHIPTYPE(request, id):
    membershiptype = MembershipType.objects.get(id = id)
    membershiptype.delete()
    messages.success(request, 'MembershipType successfully deleted')
    return redirect('view_membershiptype')

def EDIT_MEMBERSHIPTYPE(request, id):
    membershiptype = MembershipType.objects.filter(id = id)
    
    context = {
        'membershiptype':membershiptype,
    }
    
    return render(request,'hoo/edit_membershiptype.html', context)

def UPDATE_MEMBERSHIPTYPE(request):
    if request.method == "POST":
        id = request.POST.get('membershiptype_id')
        membershiptype_name = request.POST.get('membershiptype_name')
        membershiptype_info = request.POST.get('membershiptype_info')
        membershiptype_price = request.POST.get('membershiptype_price')
        # print(program)
        
        membershiptype = MembershipType.objects.get(id = id)
        membershiptype.name = membershiptype_name
        membershiptype.info = membershiptype_info
        membershiptype.price = membershiptype_price
        
        membershiptype.save()
        
        messages.success(request, "MembershipType successfully updated")
        return redirect('view_membershiptype')
    return render(request, 'hoo/edit_membershiptype.html')


#For Membertype Modules
def VIEW_MEMBERTYPE(request):
    membertype = MemberType.objects.all()
    
    context = {
        'membertype':membertype,
    }
    # print(teacher)
    return render(request, 'hoo/view_membertype.html', context)

def ADD_MEMBERTYPE(request):
    if request.method == "POST":
        membertype_name = request.POST.get('membertype_name')
        membertype_info = request.POST.get('membertype_info')
        # print(program_name)
        
        membertype = MemberType (
            name = membertype_name,
            info = membertype_info,
            created_by_id=request.user.id  # Set the created_by_username to the current user
        )
        membertype.save()
        messages.success(request, 'MemberType successfully added!')
        return redirect('view_membertype')
    return render(request, 'hoo/add_membertype.html')

def DELETE_MEMBERTYPE(request, id):
    membertype = MemberType.objects.get(id = id)
    membertype.delete()
    messages.success(request, 'MemberType successfully deleted')
    return redirect('view_membertype')

def EDIT_MEMBERTYPE(request, id):
    membertype = MemberType.objects.filter(id = id)
    
    context = {
        'membertype':membertype,
    }
    
    return render(request,'hoo/edit_membertype.html', context)

def UPDATE_MEMBERTYPE(request):
    if request.method == "POST":
        id = request.POST.get('membertype_id')
        membertype_name = request.POST.get('membertype_name')
        membertype_info = request.POST.get('membertype_info')
        # print(program)
        
        membertype = MemberType.objects.get(id = id)
        membertype.name = membertype_name
        membertype.info = membertype_info
        
        membertype.save()
        
        messages.success(request, "MemberType successfully updated")
        return redirect('view_membertype')
    return render(request, 'hoo/edit_membertype.html')

#For Organization Modules
def VIEW_ORGANIZATION(request):
    organization = Organization.objects.all()
    
    context = {
        'organization':organization,
    }
    # print(teacher)
    return render(request, 'hoo/view_organization.html', context)

def ADD_ORGANIZATION(request):
    if request.method == "POST":
        organization_initials = request.POST.get('organization_initials')
        organization_name = request.POST.get('organization_name')
        organization_type = request.POST.get('organization_type')
        organization_logo = request.FILES.get('organization_logo')
        organization_telno = request.POST.get('organization_telno')
        # print(program_name)
        
        organization = Organization (
            initials = organization_initials,
            name = organization_name,
            type = organization_type,
            logo = organization_logo,
            telephone = organization_telno,
            # created_by_id=request.user.id  # Set the created_by_username to the current user
        )
        organization.save()
        messages.success(request, 'Organization successfully added!')
        return redirect('view_organization')
    return render(request, 'hoo/add_organization.html')

def DELETE_ORGANIZATION(request, id):
    organization = Organization.objects.get(id = id)
    organization.delete()
    messages.success(request, 'Organization successfully deleted')
    return redirect('view_organization')

def EDIT_ORGANIZATION(request, id):
    organization = Organization.objects.filter(id = id)
    organizations = Organization.objects.all()
    
    context = {
        'organization':organization,
        'organizations':organizations,
    }
    
    return render(request,'hoo/edit_organization.html', context)

def UPDATE_ORGANIZATION(request):
    if request.method == "POST":
        id = request.POST.get('organization_id')
        organization_initials = request.POST.get('organization_initials')
        organization_name = request.POST.get('organization_name')
        organization_type = request.POST.get('organization_type')
        organization_logo = request.FILES.get('organization_logo')
        organization_telno = request.POST.get('organization_telno')
        # print(program)
        
        organization = Organization.objects.get(id = id)
        organization.initials = organization_initials
        organization.name = organization_name
        organization.type = organization_type
        organization.logo = organization_logo
        organization.telephone = organization_telno
        
        organization.save()
        
        messages.success(request, "Organization successfully updated")
        return redirect('view_organization')
    return render(request, 'hoo/edit_organization.html')

#For Announcement Modules
def VIEW_ANNOUNCEMENT(request):
    announcements = Announcement.objects.all()
    
    context = {
        'announcements':announcements,
    }
    # print(teacher)
    return render(request, 'hoo/view_announcement.html', context)

def ADD_ANNOUNCEMENT(request):
    if request.method == "POST":
        announcement_title = request.POST.get('announcement_title')
        announcement_description = request.POST.get('announcement_description')
        announcement_banner = request.FILES.get('announcement_banner')
        announcement_status = request.POST.get('announcement_status')
        announcement_tags = request.POST.get('announcement_tags')
        # print(program_name)
        
        announcement = Announcement (
            title = announcement_title,
            description = announcement_description,
            banner = announcement_banner,
            status = announcement_status,
            created_by_id=request.user.id  # Set the created_by_username to the current user
        )
        announcement.save()
        
        # Handle tags
        if announcement_tags:
            tags_list = [tag.strip() for tag in announcement_tags.split(',')]
            announcement.tags.add(*tags_list)  # Add tags to the announcement
        
        
        messages.success(request, 'Announcement successfully added!')
        return redirect('view_announcement')
    return render(request, 'hoo/add_announcement.html')

def DELETE_ANNOUNCEMENT(request, id):
    announcement = Announcement.objects.get(id = id)
    announcement.delete()
    messages.success(request, 'Announcement successfully deleted')
    return redirect('view_announcement')

def EDIT_ANNOUNCEMENT(request, id):
    announcement = Announcement.objects.filter(id = id)
    announcements = Announcement.objects.all()
    # tags_string = ', '.join([str(tag) for tag in announcement.tags.all()])
    
    context = {
        'announcement':announcement,
        'announcements':announcements,
        # 'tags_string': tags_string,  # Pass the tags string to the template
    }

    return render(request,'hoo/edit_announcement.html', context)

def UPDATE_ANNOUNCEMENT(request):
    if request.method == "POST":
        id = request.POST.get('announcement_id')
        announcement_title = request.POST.get('announcement_title')
        announcement_description = request.POST.get('announcement_description')
        announcement_banner = request.FILES.get('announcement_banner')
        announcement_status = request.POST.get('announcement_status')
        announcement_tags = request.POST.get('announcement_tags')
        # print(program)
        # Convert the string value to a boolean
       
        announcement = Announcement.objects.get(id = id)
        announcement.title = announcement_title
        announcement.description = announcement_description
        announcement.banner = announcement_banner
        
        if announcement_status == '1':
            announcement.status = True
        elif announcement_status == '0':
            announcement.status = False
        else:
            # Handle case where no valid status is provided
            announcement.status = None  # or keep it unchanged
        # announcement.status = announcement_status,
        
        if announcement_tags:
            # Split the tags into a list
            tags_list = [tag.strip() for tag in announcement_tags.split(',')]
            # Clear existing tags and add new ones
            announcement.tags.set(*tags_list)  # Use set() to replace existing tags
        
        announcement.save()
        
        messages.success(request, "Announcement successfully updated")
        return redirect('view_announcement')
    return render(request, 'hoo/edit_announcement.html')

def MEMBERSHIP_REGISTRATION(request):
    schoolyear = School_Year.objects.all()
    
    context = {
        'schoolyear':schoolyear,
    }
    # print(teacher)
    return render(request, 'hoo/membership_registration.html', context)


@login_required(login_url='/')
@require_http_methods(["GET", "POST"])
def MEMBERSHIP_APPROVAL(request):
    """
    View for Membership Registration Approval.
    GET: Fetch all members and map membership status.
    POST: Update membership status using membership.member_id for uniqueness.
    Sends emails using CustomUser email linked via Member.admin.
    """
    if request.method == "POST":
        # Support AJAX bulk requests: accept JSON { ids: [...], action: 'approve'|'decline' }
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            try:
                payload = json.loads(request.body.decode('utf-8') or '{}')
            except Exception:
                return JsonResponse({'error': 'bad_payload'}, status=400)

            ids = payload.get('ids') if isinstance(payload, dict) else None
            action = payload.get('action') if isinstance(payload, dict) else None

            if not ids or not isinstance(ids, list) or action not in ('approve', 'decline'):
                return JsonResponse({'error': 'missing_ids_or_action'}, status=400)

            approved = []
            declined = []
            failed = {}
            email_failed = {}

            for mid in ids:
                try:
                    member = Member.objects.get(id=mid)
                except Member.DoesNotExist:
                    failed[str(mid)] = 'not_found'
                    continue

                user = getattr(member, 'admin', None)

                # Find membership record (prefer pending)
                membership_qs = Membership.objects.filter(member_id=member.id).order_by('-id')
                membership = membership_qs.filter(status__iexact='PENDING').first() or membership_qs.first()
                if not membership:
                    failed[str(mid)] = 'no_membership'
                    continue

                try:
                    if action == 'approve':
                        membership.status = 'APPROVED'
                        membership.save()
                        if user:
                            user.is_active = 1
                            user.save()

                        # For new memberships generate password and email; for renewals just notify
                        try:
                            if getattr(membership, 'membertype_id', None) != 2:
                                password = get_random_string(length=10)
                                if user:
                                    user.set_password(password)
                                    user.save()
                                from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                                try:
                                    send_mail(
                                        subject='Membership Approved',
                                        message=(
                                            f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                            f"Your membership has been approved.\n"
                                            f"Here are your login credentials:\n\n"
                                            f"Username: {getattr(user, 'username', '')}\n"
                                            f"Password: {password}\n\n"
                                            f"Please login and change your password as soon as possible."
                                        ),
                                        from_email=from_email,
                                        recipient_list=[getattr(user, 'email', '')],
                                        fail_silently=False,
                                    )
                                except Exception as e:
                                    email_failed[str(mid)] = str(e)
                            else:
                                from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                                try:
                                    send_mail(
                                        subject='Membership Renewal Approved',
                                        message=(
                                            f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                            f"Your membership renewal has been approved.\n\n"
                                            f"Thank you."
                                        ),
                                        from_email=from_email,
                                        recipient_list=[getattr(user, 'email', '')],
                                        fail_silently=False,
                                    )
                                except Exception as e:
                                    email_failed[str(mid)] = str(e)
                        except Exception:
                            pass
                        approved.append(mid)

                    elif action == 'decline':
                        membership.status = 'DECLINED'
                        membership.save()
                        try:
                            from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                            try:
                                send_mail(
                                    subject='Membership Application Declined',
                                    message=(
                                        f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                        "We regret to inform you that your membership application has been declined.\n\n"
                                        "If you believe this is a mistake or have questions, please contact the administration.\n\n"
                                        "Regards,\nPSITE President/Officer"
                                    ),
                                    from_email=from_email,
                                    recipient_list=[getattr(user, 'email', '')],
                                    fail_silently=False,
                                )
                            except Exception as e:
                                email_failed[str(mid)] = str(e)
                        except Exception:
                            pass
                        declined.append(mid)

                except Exception as e:
                    failed[str(mid)] = str(e)

            # Add single summary messages for the HOO page reload
            try:
                if approved:
                    messages.success(request, f"{len(approved)} membership(s) approved.")
                if declined:
                    messages.warning(request, f"{len(declined)} membership(s) declined.")
                if failed:
                    messages.error(request, f"{len(failed)} membership(s) failed to process.")
                if email_failed:
                    messages.warning(request, f"{len(email_failed)} email(s) failed to send.")
            except Exception:
                pass

            return JsonResponse({'success': True, 'approved': approved, 'declined': declined, 'failed': failed, 'email_failed': email_failed})

        # Non-AJAX single-member form POST (existing behavior)
        member_id = request.POST.get("member_id")
        action = request.POST.get("action")  # "approve" or "decline"

        # Find Member
        member = get_object_or_404(Member, id=member_id)
        # Get CustomUser details for name/email
        user = member.admin

        # Get related Membership
        # A member may have multiple Membership records (new + renew). Select the most
        # recent one to act upon. Prefer pending records if present; otherwise fall back
        # to the latest by id.
        membership_qs = Membership.objects.filter(member_id=member.id).order_by('-id')
        # Prefer a PENDING membership if one exists
        membership = membership_qs.filter(status__iexact='PENDING').first()
        if not membership:
            # Fall back to the latest membership record for this member
            membership = membership_qs.first()

        if not membership:
            messages.error(request, f"No membership record found for {user.first_name} {user.last_name}.")
            return redirect("membership_approval")

        if action == "approve":
            # Approve the membership record
            membership.status = "APPROVED"
            membership.save()

            # Activate the user account
            user.is_active = 1
            user.save()

            # If this is a NEW membership (membertype != 2) generate credentials and email them.
            # For RENEWALS (membertype_id == 2) do NOT create a new password — just notify the user.
            if getattr(membership, 'membertype_id', None) != 2:
                # ✅ Generate new password for NEW registrations
                password = get_random_string(length=10)
                user.set_password(password)
                user.save()

                # ✅ Send approval email with username and password
                send_mail(
                    subject="Membership Approved",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"Your membership has been approved.\n"
                        f"Here are your login credentials:\n\n"
                        f"Username: {user.username}\n"
                        f"Password: {password}\n\n"
                        f"Please login and change your password as soon as possible."
                    ),
                    from_email="yourgmail@gmail.com",  # Replace with your EMAIL_HOST_USER
                    recipient_list=[user.email],
                    fail_silently=False,
                )

                messages.success(
                    request,
                    f"Membership for {user.first_name} {user.last_name} approved, user activated, and email with credentials sent."
                )
            else:
                # Renewal: do not generate a new password. Send a simple approval notification.
                send_mail(
                    subject="Membership Renewal Approved",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"Your membership renewal has been approved.\n"
                        f"No changes were made to your login credentials.\n\n"
                        f"Thank you."
                    ),
                    from_email="yourgmail@gmail.com",  # Replace with your EMAIL_HOST_USER
                    recipient_list=[user.email],
                    fail_silently=False,
                )

                messages.success(
                    request,
                    f"Membership renewal for {user.first_name} {user.last_name} approved; user retains existing credentials."
                )

        elif action == "decline":
            membership.status = "DECLINED"
            membership.save()

            # Send decline notification email to the user
            from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
            try:
                send_mail(
                    subject="Membership Application Declined",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        "We regret to inform you that your membership application has been declined.\n\n"
                        "If you believe this is a mistake or have questions, please contact the administration.\n\n"
                        "Regards,\nPSITE President/Officer"
                    ),
                    from_email=from_email,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
            except Exception as e:
                messages.error(request, f"Failed to send decline email: {e}")

            messages.warning(
                request, f"Membership for {user.first_name} {user.last_name} declined."
            )
        else:
            messages.error(request, "Invalid action.")

        return redirect("membership_approval")

    # --- GET logic ---
    # Only consider memberships that are either NEW (membertype_id=1) or RENEW (membertype_id=2)
    memberships = (
        Membership.objects.filter(membertype_id__in=[1, 2])
        .order_by('-id')
        .only('member_id', 'status', 'proof_of_payment', 'membertype_id')
    )

    # Build maps using the latest membership per member (ordered by -id above)
    membership_status_map = {}
    membership_payment_map = {}
    membership_type_map = {}
    for m in memberships:
        if m.member_id not in membership_status_map:
            membership_status_map[m.member_id] = m.status
            membership_payment_map[m.member_id] = m.proof_of_payment
            membership_type_map[m.member_id] = getattr(m, 'membertype_id', None)

    # Get only members who have a relevant membership (new or renew)
    member_ids = list(membership_status_map.keys())
    members = Member.objects.select_related('admin', 'organization', 'membershiptype').filter(id__in=member_ids)

    # Add dynamic attributes for display using the maps
    for member in members:
        member.status = membership_status_map.get(member.id, 'PENDING')
        member.proof_of_payment = membership_payment_map.get(member.id, None)
        member.latest_membertype_id = membership_type_map.get(member.id)

    context = {
        'schoolyear': School_Year.objects.all(),
        'users': CustomUser.objects.all(),
        'members': members,
        'membership_status_map': membership_status_map,
        'memberships': memberships,
        'membership_types': MembershipType.objects.all(),
        'member_types': MemberType.objects.all(),
        'organization': Organization.objects.all(),
    }

    return render(request, 'hoo/membership_approval.html', context)



def VIEWALL_EVENT(request):
    # Fetch all events with related school year to reduce queries
    events = Event.objects.select_related('school_year').all()
    return render(request, 'hoo/viewall_event.html', {'events': events})


@login_required(login_url='/')
def EVENT_INVITATIONS(request):
    """Render the member invitations/list table used by `event_invitations.html`.

    Provides `members` queryset with related `admin` and `organization` to
    avoid N+1 queries in the template.
    """
    # Allow filtering by membership status (based on the latest Membership row per member)
    status_filter = request.GET.get('status_filter', 'all')  # values: 'all', 'active', 'inactive'

    base_qs = Member.objects.select_related('admin', 'organization')

    if status_filter in ('active', 'inactive'):
        # Find the latest membership id per member, then pick those with APPROVED/DECLINED
        latest = Membership.objects.values('member_id').annotate(latest_id=Max('id'))
        latest_ids = [r['latest_id'] for r in latest if r.get('latest_id')]

        if latest_ids:
            if status_filter == 'active':
                member_ids = list(Membership.objects.filter(id__in=latest_ids, status__iexact='APPROVED').values_list('member_id', flat=True))
            else:
                member_ids = list(Membership.objects.filter(id__in=latest_ids, status__iexact='DECLINED').values_list('member_id', flat=True))

            members = base_qs.filter(id__in=member_ids)
        else:
            members = Member.objects.none()
    else:
        members = base_qs.all()

    context = {
        'members': members,
        'status_filter': status_filter,
    }
    return render(request, 'hoo/event_invitations.html', context)


@login_required(login_url='/')
@require_http_methods(["POST"])
def SEND_INVITATIONS(request):
    """Send invitation emails to selected members.

    Expects POST with:
      - member_ids: repeated member id values
      - subject: email subject
      - body: email body

    Includes active event title and banner URL in the message. Attaches the
    banner file if available on disk.
    """
    member_ids = request.POST.getlist('member_ids')
    subject = request.POST.get('subject', '').strip()
    body = request.POST.get('body', '').strip()

    if not member_ids:
        messages.error(request, 'No members selected for invitations.')
        return redirect('hoo_event_invitations')

    if not subject or not body:
        messages.error(request, 'Subject and message body are required.')
        return redirect('hoo_event_invitations')

    # Resolve recipients
    members = Member.objects.select_related('admin').filter(id__in=member_ids)
    recipients = [m.admin.email for m in members if getattr(m, 'admin', None) and getattr(m.admin, 'email', None)]

    if not recipients:
        messages.error(request, 'Selected members do not have email addresses.')
        return redirect('hoo_event_invitations')

    # Get active event (if any)
    event = Event.objects.filter(status='active').order_by('-date').first()
    event_title = event.title if event else ''
    banner_url = ''
    banner_path = None
    qr_url = ''
    qr_path = None
    if event and getattr(event, 'banner', None):
        try:
            banner_url = request.build_absolute_uri(event.banner.url)
        except Exception:
            banner_url = ''
        # try to attach if file path exists
        try:
            banner_path = event.banner.path
        except Exception:
            banner_path = None
    # Try to get qr path/url
    if event and getattr(event, 'qr_code', None):
        try:
            qr_url = request.build_absolute_uri(event.qr_code.url)
        except Exception:
            qr_url = ''
        try:
            qr_path = event.qr_code.path
        except Exception:
            qr_path = None

    # Compose message: include provided body then append event info
    message = f"{body}\n\n"
    if event_title:
        message += f"Event: {event_title}\n"
    if banner_url:
        message += f"Banner: {banner_url}\n"

    # Determine from-email
    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', None) or getattr(settings, 'EMAIL_HOST_USER', None)
    try:
        # Prepare plain text and HTML content. HTML will embed the banner inline.
        text_content = message
        # If there's an inline banner, reference it via cid:bannerimg and display it smaller
        # Build HTML content: optional inline banner, then body, then optional QR image
        parts = []
        if banner_path and os.path.exists(banner_path):
            parts.append(
                f"<img src=\"cid:bannerimg\" alt=\"Event Banner\" "
                f"style=\"display:block;width:80px;height:80px;object-fit:cover;margin:0 0 8px 0;\">"
            )
        parts.append(f"<div>{body.replace('\n','<br>')}</div>")
        if qr_path and os.path.exists(qr_path):
            parts.append(
                f"<div style=\"margin-top:12px;\">"
                f"<p style=\"margin:0 0 6px 0;\"><strong>Scan QR to register:</strong></p>"
                f"<img src=\"cid:qrimg\" alt=\"Event QR\" style=\"width:200px;height:200px;object-fit:contain;\">"
                f"</div>"
            )
        html_content = f"<div>{''.join(parts)}</div>"

        msg = EmailMultiAlternatives(subject=subject, body=text_content, from_email=from_email, to=[from_email], bcc=recipients)
        msg.attach_alternative(html_content, "text/html")

        # Attach banner as inline image (MIMEImage) so it appears in the body instead of a link
        if banner_path and os.path.exists(banner_path):
            try:
                with open(banner_path, 'rb') as f:
                    img_data = f.read()
                mime_img = MIMEImage(img_data)
                mime_img.add_header('Content-ID', '<bannerimg>')
                mime_img.add_header('Content-Disposition', 'inline; filename="%s"' % os.path.basename(banner_path))
                msg.attach(mime_img)
            except Exception:
                # ignore image attach failures
                pass

        # Attach QR image inline under the body if present
        if qr_path and os.path.exists(qr_path):
            try:
                with open(qr_path, 'rb') as f:
                    qr_data = f.read()
                mime_qr = MIMEImage(qr_data)
                mime_qr.add_header('Content-ID', '<qrimg>')
                mime_qr.add_header('Content-Disposition', 'inline; filename="%s"' % os.path.basename(qr_path))
                msg.attach(mime_qr)
            except Exception:
                pass

        # Attach uploaded PDF(s) (if provided) — support multiple files
        uploaded_files = request.FILES.getlist('attachment')
        for uploaded in uploaded_files:
            try:
                content = uploaded.read()
                content_type = getattr(uploaded, 'content_type', 'application/pdf')
                msg.attach(uploaded.name, content, content_type)
            except Exception:
                # ignore attachment failures for individual files
                continue

        msg.send(fail_silently=False)
        messages.success(request, f'Invitations sent to {len(recipients)} recipients.')
    except Exception as e:
        messages.error(request, f'Failed to send invitations: {str(e)}')

    return redirect('hoo_event_invitations')



def GET_EVENT_JSON(request, id):
    """Return event details as JSON for the View modal."""
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            event = Event.objects.get(pk=id)
            data = {
                "id": event.id,
                "title": event.title,
                "theme": event.theme,
                "date": event.date.strftime("%Y-%m-%d") if event.date else "",
                "location": event.location,
                "max_attendees": event.max_attendees,
                "registration_fee": str(event.registration_fee) if event.registration_fee else "",
                "school_year": f"{event.school_year.sy_start} - {event.school_year.sy_end}" if event.school_year else "N/A",
            }
            return JsonResponse(data)
        except Event.DoesNotExist:
            raise Http404("Event not found")
    else:
        raise Http404("Invalid request")



def ADD_EVENT(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        theme = request.POST.get('theme')
        date = request.POST.get('date')
        location = request.POST.get('location')
        max_attendees = request.POST.get('max_attendees')
        registration_fee = request.POST.get('registration_fee')
        chair_id = request.POST.get('chair')
        co_chair_id = request.POST.get('co_chair')
        evaluation_link = request.POST.get('evaluation_link')

        banner = request.FILES.get('banner')
        # Selected tag id from form (optional)
        tag_id = request.POST.get('tag')

        # ✅ Get active school year
        try:
            active_schoolyear = School_Year.objects.get(status=1)
        except School_Year.DoesNotExist:
            messages.error(request, "No active school year found. Please activate a school year first.")
            return redirect('add_event')

        # ✅ Deactivate previous active events
        Event.objects.filter(status='active').update(status='inactive')

        # ✅ Create and save new event
        # ensure numeric fields are cast
        try:
            max_attendees_val = int(max_attendees) if max_attendees not in (None, '') else 0
        except Exception:
            max_attendees_val = 0
        try:
            registration_fee_val = float(registration_fee) if registration_fee not in (None, '') else 0.0
        except Exception:
            registration_fee_val = 0.0

        event = Event(
            title=title,
            theme=theme,
            date=date,
            location=location,
            max_attendees=max_attendees_val,
            registration_fee=registration_fee_val,
            chair_id=chair_id,
            co_chair_id=co_chair_id,
            # registration_link will be generated after saving (absolute URL to registration page)
            evaluation_link=evaluation_link,
            banner=banner,
            created_by=request.user,
            created_at=timezone.now(),
            updated_at=timezone.now(),
            school_year=active_schoolyear,
            status='active',
            available_slots=max_attendees_val,
            tags_id=tag_id if tag_id not in (None, '') else None
        )
        event.save()
        # Build absolute registration link and save it on the event
        try:
            reg_path = reverse('registration_event') + f'?event_id={event.id}'
            registration_url = request.build_absolute_uri(reg_path)
            event.registration_link = registration_url
            event.save()
        except Exception as e:
            print('Failed to set registration_link for event', event.id, e)

        # Auto-generate QR code image encoding the registration URL and save path to event.qr_code
        try:
            qr_data = event.registration_link if event.registration_link else f"EVENT:{event.id}"
            img = qrcode.make(qr_data)
            qr_dir = os.path.join(settings.MEDIA_ROOT, 'qr_codes')
            os.makedirs(qr_dir, exist_ok=True)
            filename = f"event_{event.id}_qr.png"
            dest_path = os.path.join(qr_dir, filename)
            img.save(dest_path)
            # Save relative media path on the model (use forward slashes)
            event.qr_code = os.path.join('qr_codes', filename).replace('\\', '/')
            event.save()
        except Exception as e:
            print('Failed to generate QR code for event', event.id, e)

        # Handle optional bulk registration template upload
        bulk_template = request.FILES.get('bulk_template')
        if bulk_template:
            try:
                # Ensure directory exists under MEDIA_ROOT/bulk_template
                dest_dir = os.path.join(settings.MEDIA_ROOT, 'bulk_template')
                os.makedirs(dest_dir, exist_ok=True)
                # Create a safe filename with timestamp to avoid collisions
                timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
                filename = f"{timestamp}_{bulk_template.name}"
                dest_path = os.path.join(dest_dir, filename)
                # Write file to disk
                with open(dest_path, 'wb+') as dst:
                    for chunk in bulk_template.chunks():
                        dst.write(chunk)
                # Save relative path to event.template_path
                event.template_path = os.path.join('bulk_template', filename).replace('\\', '/')
                event.save()
            except Exception as e:
                # log but don't break the flow
                print('Failed to save bulk_template:', e)

        messages.success(
            request,
            f'Event added successfully for cycle {active_schoolyear.sy_start.year} - {active_schoolyear.sy_end.year}!'
        )
        return redirect('viewall_event')

    # Fetch dropdown data
    members = Member.objects.all()
    officertypes = OfficerType.objects.all()
    custom_users = CustomUser.objects.filter(user_type__in=[1, 2], is_superuser=0)
    tags = Tags.objects.all()

    # ✅ Map officertype names to each custom user
    for user in custom_users:
        try:
            member = members.get(admin_id=user.id)
            officer_type = officertypes.get(id=member.officertype_id)
            user.officertype_name = officer_type.name
        except (Member.DoesNotExist, OfficerType.DoesNotExist):
            user.officertype_name = "N/A"

    # ✅ Active school year for template
    active_schoolyear = School_Year.objects.filter(status=1).first()

    return render(request, 'hoo/add_event.html', {
        'members': members,
        'custom_users': custom_users,
        'officertypes': officertypes,
        'active_schoolyear': active_schoolyear,
        'tags': tags
    })


def DELETE_EVENT(request, id):
    event = Event.objects.get(id = id)
    event.delete()
    messages.success(request, 'event successfully deleted')
    return redirect('viewall_event')



def EDIT_EVENT(request, id):
    event = get_object_or_404(Event, id=id)

    if request.method == 'POST':
        event.title = request.POST.get('title')
        event.theme = request.POST.get('theme')
        event.date = request.POST.get('date')
        event.location = request.POST.get('location')
        event.max_attendees = request.POST.get('max_attendees')
        event.registration_fee = request.POST.get('registration_fee')

        if 'banner' in request.FILES:
            event.banner = request.FILES['banner']

        event.save()
        messages.success(request, f'Event "{event.title}" updated successfully!')
        return redirect('viewall_event')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'id': event.id,
            'title': event.title,
            'theme': event.theme,
            'date': event.date.strftime('%Y-%m-%d'),
            'location': event.location,
            'max_attendees': event.max_attendees,
            'registration_fee': event.registration_fee,
        })

    return redirect('viewall_event')

@login_required(login_url='/')
def ATTENDANCE_EVENT(request):
    """Render attendance page for the currently active event.

        Builds an `attendees` list with dictionaries containing:
            - id: registration id
            - first_name, last_name: from CustomUser
            - school: from Member.organization.name (if any)
            - present: current `Member_Event_Registration.is_present` (True/False) or None

    The template receives `attendees` and `event` in context.
    """
    # Get the active event (most recent by date)
    active_event = Event.objects.filter(status__in=['active', 'full']).order_by('-date').first()

    attendees = []
    if active_event:
        # Get registrations for the active event (registered status)
        # Use `member_id` relationship (FK to Member) and related admin/user/org
        regs = Member_Event_Registration.objects.filter(event=active_event, status='registered').select_related('member_id__admin', 'member_id__organization')
        for reg in regs:
            member_obj = getattr(reg, 'member_id', None)
            user = getattr(member_obj, 'admin', None) if member_obj else None
            first = user.first_name if user else ''
            last = user.last_name if user else ''

            # Member's organization name
            school_name = ''
            try:
                if member_obj and getattr(member_obj, 'organization', None):
                    school_name = member_obj.organization.name
            except Exception:
                school_name = ''
                school_name = ''

            # Use the registration's `is_present` flag as the current attendance status
            # `is_present` is expected to be 1 (present), 0 (absent), or maybe None
            present = None
            try:
                present = True if getattr(reg, 'is_present', None) == 1 else (False if getattr(reg, 'is_present', None) == 0 else None)
            except Exception:
                present = None

            attendees.append({
                'id': reg.id,
                'first_name': first,
                'last_name': last,
                'school': school_name,
                'present': present,
            })

    context = {
        'attendees': attendees,
        'event': active_event,
    }
    return render(request, 'hoo/attendance_event.html', context)


@login_required(login_url='/')
@require_http_methods(["POST"])
def ATTENDANCE_TOGGLE(request):
    """AJAX endpoint to toggle attendance for a registration.

    Expects JSON payload: { id: <member_event_registration_id>, present: <true|false> }
    Updates Member_Event_Registration.is_present to 1 when present is truthy, otherwise 0.
    Returns JSON { success: True, is_present: 1|0 } on success.
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
        reg_id = payload.get('id')
        present = payload.get('present')
    except Exception:
        return JsonResponse({'error': 'Bad payload'}, status=400)

    if reg_id is None:
        return JsonResponse({'error': 'missing_id'}, status=400)

    # Try Bulk_Event_Reg first
    try:
        bulk = Bulk_Event_Reg.objects.filter(id=int(reg_id)).first()
    except Exception:
        bulk = None

    if bulk:
        try:
            bulk.is_present = 1 if bool(present) else 0
            bulk.save()
        except Exception as e:
            return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

        try:
            if bulk.is_present:
                messages.success(request, f"Marked {bulk.first_name} {bulk.last_name} as Present.")
            else:
                messages.warning(request, f"Marked {bulk.first_name} {bulk.last_name} as Absent.")
        except Exception:
            pass

        return JsonResponse({'success': True, 'id': bulk.id, 'is_present': bulk.is_present})

    # Fallback to Member_Event_Registration
    try:
        # include member->admin so we can show the member's user name in messages
        reg = Member_Event_Registration.objects.select_related('member_id__admin').get(id=int(reg_id))
    except (Member_Event_Registration.DoesNotExist, ValueError):
        return JsonResponse({'error': 'Registration not found'}, status=404)

    try:
        reg.is_present = 1 if bool(present) else 0
        reg.save()
    except Exception as e:
        return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if reg.is_present:
            if user:
                messages.success(request, f"Marked {user.first_name} {user.last_name} as Present.")
            else:
                messages.success(request, "Marked registration as Present.")
        else:
            if user:
                messages.warning(request, f"Marked {user.first_name} {user.last_name} as Absent.")
            else:
                messages.warning(request, "Marked registration as Absent.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'id': reg.id, 'is_present': reg.is_present})


def VIEWALL_BULK_REG(request):
    # Provide list of events for filtering and pick default selected event
    events = Event.objects.all().order_by('-date')
    active_event = Event.objects.filter(status='active').order_by('-date').first()

    # Determine selected event: GET param `event` takes precedence, else active_event
    selected_event_id = None
    try:
        sel = request.GET.get('event')
        if sel and str(sel).isdigit():
            selected_event_id = int(sel)
        elif active_event:
            selected_event_id = active_event.id
    except Exception:
        selected_event_id = active_event.id if active_event else None

    # Build `bulk_regs` from Member_Event_Registration entries, optionally filtered by event
    regs_qs = Member_Event_Registration.objects.select_related('member_id__organization', 'member_id__admin')
    if selected_event_id:
        regs_qs = regs_qs.filter(event_id=selected_event_id)

    bulk_regs = []
    for reg in regs_qs.all():
        member_obj = getattr(reg, 'member_id', None)
        if not member_obj:
            continue

        user = getattr(member_obj, 'admin', None)

        bulk_regs.append({
            'id': reg.id,
            'last_name': (getattr(user, 'last_name', '') or ''),
            'first_name': (getattr(user, 'first_name', '') or ''),
            # keep the Member instance so template can access registered_by.organization.name
            'registered_by': member_obj,
            # default attending_as to Professor as requested
            'attending_as': 'Professor',
            'is_approved': bool(getattr(reg, 'is_approved', False)),
            'is_present': bool(getattr(reg, 'is_present', False)),
            'event_id': getattr(reg, 'event_id', None),
            'event_title': getattr(reg, 'event', None) and getattr(reg.event, 'title', '') or '',
        })

    context = {
        'bulk_regs': bulk_regs,
        'events': events,
        'selected_event_id': selected_event_id,
    }
    return render(request, 'hoo/view_bulk_reg.html', context)


def GET_BULK_BY_MEMBER(request, member_id):
    """Return Bulk_Event_Reg entries for a given Member.id as JSON.

    Response: { "bulk_regs": [ {id, last_name, first_name, attending_as}, ... ] }
    """
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.method == 'GET':
        regs_qs = Bulk_Event_Reg.objects.filter(registered_by_id=member_id).values('id', 'last_name', 'first_name', 'attending_as')
        regs = list(regs_qs)

        member = Member.objects.filter(id=member_id).select_related('admin', 'organization').first()
        owner_name = ''
        if member:
            # Prefer organization name (school) as the owner label; fall back to admin full name
            org = getattr(member, 'organization', None)
            if org and getattr(org, 'name', None):
                owner_name = org.name
            elif getattr(member, 'admin', None):
                owner_name = f"{getattr(member.admin, 'first_name', '')} {getattr(member.admin, 'last_name', '')}".strip()

        return JsonResponse({'bulk_regs': regs, 'member_name': owner_name})
    raise Http404("Invalid request")


def APPROVE_BULK_EVENT_REG(bulk_id):
    """Approve a single Bulk_Event_Reg entry and decrement the related Event.available_slots.

    Returns (True, id) on success, (False, reason) on failure.
    """
    try:
        bulk = Bulk_Event_Reg.objects.select_related('registered_by').get(id=bulk_id)
    except Bulk_Event_Reg.DoesNotExist:
        return False, 'not_found'

    # determine related event id (support both event_id field or event relation)
    event_id = getattr(bulk, 'event_id', None) or (getattr(bulk, 'event', None) and getattr(bulk.event, 'id', None))

    try:
        with transaction.atomic():
            # if already approved, return success
            if getattr(bulk, 'is_approved', False):
                return True, bulk.id

            bulk.is_approved = True
            bulk.save()

            if event_id:
                Event.objects.filter(id=event_id).update(available_slots=F('available_slots') - 1)
                # clamp non-negative
                try:
                    ev = Event.objects.get(id=event_id)
                    if ev.available_slots is None or ev.available_slots < 0:
                        ev.available_slots = 0
                        ev.save()
                except Exception:
                    pass

        return True, bulk.id
    except Exception as e:
        return False, str(e)


def APPROVE_BULK_EVENT_REGS(bulk_ids):
    """Approve multiple Bulk_Event_Reg ids. Returns dict with successes and failures."""
    results = {'success': [], 'failed': {}}
    for bid in bulk_ids:
        ok, info = APPROVE_BULK_EVENT_REG(bid)
        if ok:
            results['success'].append(info)
        else:
            results['failed'][str(bid)] = info
    return results


@login_required(login_url='/')
@require_http_methods(["POST"])
def DECLINE_BULK_EVENT_REG(request, reg_id):
    """Decline a single Bulk_Event_Reg entry (AJAX).

    Sets `is_approved=False` and returns JSON.
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        bulk = Bulk_Event_Reg.objects.select_related('event').get(id=reg_id)
    except Bulk_Event_Reg.DoesNotExist:
        return JsonResponse({'error': 'not_found'}, status=404)

    try:
        with transaction.atomic():
            was_approved = bool(getattr(bulk, 'is_approved', False))
            bulk.is_approved = False
            bulk.save()

            # If this registration was previously approved, restore an available slot
            if was_approved:
                event_id = getattr(bulk, 'event_id', None) or (getattr(bulk, 'event', None) and getattr(bulk.event, 'id', None))
                if event_id:
                    Event.objects.filter(id=event_id).update(available_slots=F('available_slots') + 1)
                    # clamp to max_attendees if present and ensure no None
                    try:
                        ev = Event.objects.get(id=event_id)
                        if ev.available_slots is None:
                            ev.available_slots = 0
                        if getattr(ev, 'max_attendees', None) is not None and ev.available_slots > ev.max_attendees:
                            ev.available_slots = ev.max_attendees
                        ev.save()
                    except Exception:
                        pass
    except Exception as e:
        return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

    try:
        messages.warning(request, f"Bulk registration for {bulk.first_name} {bulk.last_name} declined.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'id': bulk.id})



@login_required(login_url='/')
@require_http_methods(["POST"])
def APPROVE_BULK_EVENT_REG_VIEW(request, reg_id):
    """AJAX view wrapper to approve a single Bulk_Event_Reg entry.

    Calls the helper `APPROVE_BULK_EVENT_REG` and returns JSON.
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    ok, info = APPROVE_BULK_EVENT_REG(reg_id)
    if ok:
        try:
            messages.success(request, f"Bulk registration (id={info}) approved.")
        except Exception:
            pass
        return JsonResponse({'success': True, 'id': info})
    else:
        return JsonResponse({'error': info}, status=400)


@login_required(login_url='/')
@require_http_methods(["POST"])
def APPROVE_BULK_EVENT_REGS_VIEW(request):
    """AJAX view to approve multiple Bulk_Event_Reg ids.

    Expects JSON body: { "ids": [1,2,3] }
    Returns the helper results as JSON.
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'error': 'bad_payload'}, status=400)

    ids = payload.get('ids') if isinstance(payload, dict) else None
    if not ids or not isinstance(ids, list):
        return JsonResponse({'error': 'missing_ids'}, status=400)

    results = APPROVE_BULK_EVENT_REGS(ids)
    # return both successes and failures so the client can react
    return JsonResponse({'success': True, 'results': results})


@login_required(login_url='/')
@require_http_methods(["POST"])
def APPROVE_MEMBER_EVENT_REG(request, reg_id):

    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    # Only operate on single Member_Event_Registration (no Bulk_Event_Reg handling)
    try:
        # include member->admin so message can reference the linked CustomUser
        reg = Member_Event_Registration.objects.select_related('event', 'member_id__admin').get(id=reg_id)
    except Member_Event_Registration.DoesNotExist:
        return JsonResponse({'error': 'not_found'}, status=404)

    try:
        with transaction.atomic():
            # mark approved
            reg.is_approved = True
            reg.save()

            # decrement event available_slots by 1 if event exists
            event = getattr(reg, 'event', None)
            if event and getattr(event, 'id', None):
                # use F() expression for atomic decrement
                Event.objects.filter(id=event.id).update(available_slots=F('available_slots') - 1)

                # ensure non-negative stored value
                try:
                    ev = Event.objects.get(id=event.id)
                    if ev.available_slots is None or ev.available_slots < 0:
                        ev.available_slots = 0
                        ev.save()
                except Exception:
                    # ignore failure to re-read event
                    pass
    except Exception as e:
        return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

    # Add a Django message so it appears after the page reload
    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if user:
            messages.success(request, f"Registration for {user.first_name} {user.last_name} approved.")
        else:
            messages.success(request, "Registration approved.")
    except Exception:
        pass

    # Send approval email to the registered user (non-blocking)
    try:
        event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if user and getattr(user, 'email', None):
            from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
            try:
                send_mail(
                    subject="Event Registration Approved",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"Your registration for the event \"{event_title}\" has been approved.\n\n"
                        "Please check the event details on the portal for more information.\n\n"
                        "Regards,\nPSITE President/Officer"
                    ),
                    from_email=from_email,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
            except Exception as e:
                try:
                    messages.error(request, f"Failed to send approval email: {e}")
                except Exception:
                    pass
    except Exception:
        pass

    return JsonResponse({'success': True, 'id': reg.id})


@login_required(login_url='/')
@require_http_methods(["POST"])
def APPROVE_MEMBER_EVENT_REGS_VIEW(request):
    """Approve multiple Member_Event_Registration ids and add a single summary message.

    Expects JSON body: { "ids": [1,2,3] }
    Returns JSON with success/failure counts and reloads will show the Django message.
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'error': 'bad_payload'}, status=400)

    ids = payload.get('ids') if isinstance(payload, dict) else None
    if not ids or not isinstance(ids, list):
        return JsonResponse({'error': 'missing_ids'}, status=400)

    success = []
    failed = {}
    email_failed = {}
    for rid in ids:
        try:
            reg = Member_Event_Registration.objects.select_related('event', 'member_id__admin').get(id=rid)
        except Member_Event_Registration.DoesNotExist:
            failed[str(rid)] = 'not_found'
            continue

        try:
            with transaction.atomic():
                if not getattr(reg, 'is_approved', False):
                    reg.is_approved = True
                    reg.save()
                    event = getattr(reg, 'event', None)
                    if event and getattr(event, 'id', None):
                        Event.objects.filter(id=event.id).update(available_slots=F('available_slots') - 1)
                        try:
                            ev = Event.objects.get(id=event.id)
                            if ev.available_slots is None or ev.available_slots < 0:
                                ev.available_slots = 0
                                ev.save()
                        except Exception:
                            pass
                        # Send approval email to user (non-blocking, record failures)
                        try:
                            event_title = getattr(event, 'title', 'the event')
                            member = getattr(reg, 'member_id', None)
                            user = getattr(member, 'admin', None)
                            if user and getattr(user, 'email', None):
                                from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                                try:
                                    send_mail(
                                        subject="Event Registration Approved",
                                        message=(
                                            f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                            f"Your registration for the event \"{event_title}\" has been approved.\n\n"
                                            "Please check the event details on the portal for more information.\n\n"
                                            "Regards,\nPSITE President/Officer"
                                        ),
                                        from_email=from_email,
                                        recipient_list=[user.email],
                                        fail_silently=False,
                                    )
                                except Exception as e:
                                    email_failed[str(rid)] = str(e)
                        except Exception:
                            # do not let email issues break the approval loop
                            pass
                success.append(reg.id)
        except Exception as e:
            failed[str(rid)] = str(e)

    # Add one summary message
    try:
        if success:
            messages.success(request, f"{len(success)} member(s) event registration approved.")
        if failed:
            messages.warning(request, f"{len(failed)} registration(s) failed to approve.")
        if email_failed:
            messages.warning(request, f"{len(email_failed)} approval email(s) failed to send.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'approved': success, 'failed': failed, 'email_failed': email_failed})


@login_required(login_url='/')
@require_http_methods(["POST"])
def DECLINE_MEMBER_EVENT_REGS_VIEW(request):
    """Decline multiple Member_Event_Registration ids and add a single summary message.

    Expects JSON body: { "ids": [1,2,3] }
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'error': 'bad_payload'}, status=400)

    ids = payload.get('ids') if isinstance(payload, dict) else None
    if not ids or not isinstance(ids, list):
        return JsonResponse({'error': 'missing_ids'}, status=400)

    success = []
    failed = {}
    email_failed = {}
    for rid in ids:
        try:
            reg = Member_Event_Registration.objects.select_related('event', 'member_id__admin').get(id=rid)
        except Member_Event_Registration.DoesNotExist:
            failed[str(rid)] = 'not_found'
            continue

        try:
            with transaction.atomic():
                was_approved = bool(getattr(reg, 'is_approved', False))
                reg.is_approved = False
                reg.save()

                if was_approved:
                    event = getattr(reg, 'event', None)
                    if event and getattr(event, 'id', None):
                        Event.objects.filter(id=event.id).update(available_slots=F('available_slots') + 1)
                        try:
                            ev = Event.objects.get(id=event.id)
                            if ev.available_slots is None:
                                ev.available_slots = 0
                            if getattr(ev, 'max_attendees', None) is not None and ev.available_slots > ev.max_attendees:
                                ev.available_slots = ev.max_attendees
                            ev.save()
                        except Exception:
                            pass
                        # Send decline email to user (non-blocking, record failures)
                        try:
                            event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
                            member = getattr(reg, 'member_id', None)
                            user = getattr(member, 'admin', None)
                            if user and getattr(user, 'email', None):
                                from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                                try:
                                    send_mail(
                                        subject="Event Registration Declined",
                                        message=(
                                            f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                            f"We regret to inform you that your registration for the event \"{event_title}\" has been declined.\n\n"
                                            "If you have questions, please contact the organizers via the portal.\n\n"
                                            "Regards,\nPSITE President/Officer"
                                        ),
                                        from_email=from_email,
                                        recipient_list=[user.email],
                                        fail_silently=False,
                                    )
                                except Exception as e:
                                    email_failed[str(rid)] = str(e)
                        except Exception:
                            pass
                success.append(reg.id)
        except Exception as e:
            failed[str(rid)] = str(e)

    try:
        if success:
            messages.success(request, f"{len(success)} member(s) event registration declined.")
        if failed:
            messages.warning(request, f"{len(failed)} registration(s) failed to decline.")
        if email_failed:
            messages.warning(request, f"{len(email_failed)} decline email(s) failed to send.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'declined': success, 'failed': failed, 'email_failed': email_failed})


@login_required(login_url='/')
@require_http_methods(["POST"])
def PRESENT_MEMBER_EVENT_REGS_VIEW(request):
    """Mark multiple Member_Event_Registration rows as present and email users.

    Expects JSON body: { "ids": [1,2,3] }
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'error': 'bad_payload'}, status=400)

    ids = payload.get('ids') if isinstance(payload, dict) else None
    if not ids or not isinstance(ids, list):
        return JsonResponse({'error': 'missing_ids'}, status=400)

    success = []
    failed = {}
    email_failed = {}
    for rid in ids:
        try:
            reg = Member_Event_Registration.objects.select_related('member_id__admin', 'event').get(id=rid)
        except Member_Event_Registration.DoesNotExist:
            failed[str(rid)] = 'not_found'
            continue

        try:
            with transaction.atomic():
                reg.is_present = True
                reg.save()

                # send present notification
                try:
                    member = getattr(reg, 'member_id', None)
                    user = getattr(member, 'admin', None)
                    event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
                    if user and getattr(user, 'email', None):
                        from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                        try:
                            send_mail(
                                subject="Event Attendance: Present",
                                message=(
                                    f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                    f"You have been marked present for the event \"{event_title}\".\n\n"
                                    "Thank you for attending.\n\n"
                                    "Regards,\nPSITE President/Officer"
                                ),
                                from_email=from_email,
                                recipient_list=[user.email],
                                fail_silently=False,
                            )
                        except Exception as e:
                            email_failed[str(rid)] = str(e)
                except Exception:
                    pass

                success.append(reg.id)
        except Exception as e:
            failed[str(rid)] = str(e)

    try:
        if success:
            messages.success(request, f"{len(success)} member(s) marked present.")
        if failed:
            messages.warning(request, f"{len(failed)} registration(s) failed to update present.")
        if email_failed:
            messages.warning(request, f"{len(email_failed)} present email(s) failed to send.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'presented': success, 'failed': failed, 'email_failed': email_failed})


@login_required(login_url='/')
@require_http_methods(["POST"])
def ABSENT_MEMBER_EVENT_REGS_VIEW(request):
    """Mark multiple Member_Event_Registration rows as absent and email users.

    Expects JSON body: { "ids": [1,2,3] }
    """
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return JsonResponse({'error': 'bad_payload'}, status=400)

    ids = payload.get('ids') if isinstance(payload, dict) else None
    if not ids or not isinstance(ids, list):
        return JsonResponse({'error': 'missing_ids'}, status=400)

    success = []
    failed = {}
    email_failed = {}
    for rid in ids:
        try:
            reg = Member_Event_Registration.objects.select_related('member_id__admin', 'event').get(id=rid)
        except Member_Event_Registration.DoesNotExist:
            failed[str(rid)] = 'not_found'
            continue

        try:
            with transaction.atomic():
                reg.is_present = False
                reg.save()

                # send absent notification
                try:
                    member = getattr(reg, 'member_id', None)
                    user = getattr(member, 'admin', None)
                    event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
                    if user and getattr(user, 'email', None):
                        from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
                        try:
                            send_mail(
                                subject="Event Attendance: Marked Absent",
                                message=(
                                    f"Dear {getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')},\n\n"
                                    f"You have been marked absent for the event \"{event_title}\".\n\n"
                                    "If this is incorrect, please contact the organizers.\n\n"
                                    "Regards,\nPSITE President/Officer"
                                ),
                                from_email=from_email,
                                recipient_list=[user.email],
                                fail_silently=False,
                            )
                        except Exception as e:
                            email_failed[str(rid)] = str(e)
                except Exception:
                    pass

                success.append(reg.id)
        except Exception as e:
            failed[str(rid)] = str(e)

    try:
        if success:
            messages.success(request, f"{len(success)} member(s) marked absent.")
        if failed:
            messages.warning(request, f"{len(failed)} registration(s) failed to update absent.")
        if email_failed:
            messages.warning(request, f"{len(email_failed)} absent email(s) failed to send.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'absented': success, 'failed': failed, 'email_failed': email_failed})

@login_required(login_url='/')
@require_http_methods(["POST"])
def DECLINE_MEMBER_EVENT_REG(request, reg_id):
    """Mark a Member_Event_Registration as not approved (is_approved=False)."""
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
    # Only operate on Member_Event_Registration (no Bulk_Event_Reg handling here)
    try:
        reg = Member_Event_Registration.objects.select_related('event', 'member_id__admin').get(id=reg_id)
    except Member_Event_Registration.DoesNotExist:
        return JsonResponse({'error': 'not_found'}, status=404)

    try:
        with transaction.atomic():
            was_approved = bool(getattr(reg, 'is_approved', False))
            reg.is_approved = False
            reg.save()

            # restore event slot if it was approved before
            if was_approved:
                event = getattr(reg, 'event', None)
                if event and getattr(event, 'id', None):
                    Event.objects.filter(id=event.id).update(available_slots=F('available_slots') + 1)
                    try:
                        ev = Event.objects.get(id=event.id)
                        if ev.available_slots is None:
                            ev.available_slots = 0
                        if getattr(ev, 'max_attendees', None) is not None and ev.available_slots > ev.max_attendees:
                            ev.available_slots = ev.max_attendees
                        ev.save()
                    except Exception:
                        pass
    except Exception as e:
        return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

    # Add a Django message for decline
    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if user:
            messages.warning(request, f"Registration for {user.first_name} {user.last_name} declined.")
        else:
            messages.warning(request, "Registration declined.")
    except Exception:
        pass

    # Send decline email to the registered user (non-blocking)
    try:
        event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if user and getattr(user, 'email', None):
            from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
            try:
                send_mail(
                    subject="Event Registration Declined",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"We regret to inform you that your registration for the event \"{event_title}\" has been declined.\n\n"
                        "If you believe this is an error or have questions, please contact the administration.\n\n"
                        "Regards,\nPSITE President/Officer"
                    ),
                    from_email=from_email,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
            except Exception as e:
                try:
                    messages.error(request, f"Failed to send decline email: {e}")
                except Exception:
                    pass
    except Exception:
        pass

    return JsonResponse({'success': True, 'id': reg.id})


@login_required(login_url='/')
@require_http_methods(["POST"])
def PRESENT_MEMBER_EVENT_REG(request, reg_id):
    """Mark a single Member_Event_Registration as present and email the user."""
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        reg = Member_Event_Registration.objects.select_related('member_id__admin', 'event').get(id=reg_id)
    except Member_Event_Registration.DoesNotExist:
        return JsonResponse({'error': 'not_found'}, status=404)

    try:
        with transaction.atomic():
            reg.is_present = True
            reg.save()
    except Exception as e:
        return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
        if user and getattr(user, 'email', None):
            from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
            try:
                send_mail(
                    subject="Event Attendance: Present",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"You have been marked present for the event \"{event_title}\".\n\n"
                        "Thank you for attending.\n\n"
                        "Regards,\nPSITE President/Officer"
                    ),
                    from_email=from_email,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
            except Exception as e:
                try:
                    messages.warning(request, f"Failed to send present email: {e}")
                except Exception:
                    pass
    except Exception:
        pass

    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if user:
            messages.success(request, f"Marked {user.first_name} {user.last_name} present.")
        else:
            messages.success(request, "Marked attendee present.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'id': reg.id})


@login_required(login_url='/')
@require_http_methods(["POST"])
def ABSENT_MEMBER_EVENT_REG(request, reg_id):
    """Mark a single Member_Event_Registration as absent and email the user."""
    if request.headers.get('x-requested-with') != 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        reg = Member_Event_Registration.objects.select_related('member_id__admin', 'event').get(id=reg_id)
    except Member_Event_Registration.DoesNotExist:
        return JsonResponse({'error': 'not_found'}, status=404)

    try:
        with transaction.atomic():
            reg.is_present = False
            reg.save()
    except Exception as e:
        return JsonResponse({'error': 'save_failed', 'message': str(e)}, status=500)

    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        event_title = getattr(getattr(reg, 'event', None), 'title', 'the event')
        if user and getattr(user, 'email', None):
            from_email = getattr(settings, 'EMAIL_HOST_USER', 'yourgmail@gmail.com')
            try:
                send_mail(
                    subject="Event Attendance: Marked Absent",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"You have been marked absent for the event \"{event_title}\".\n\n"
                        "If this is incorrect, please contact the organizers.\n\n"
                        "Regards,\nPSITE President/Officer"
                    ),
                    from_email=from_email,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
            except Exception as e:
                try:
                    messages.warning(request, f"Failed to send absent email: {e}")
                except Exception:
                    pass
    except Exception:
        pass

    try:
        member = getattr(reg, 'member_id', None)
        user = getattr(member, 'admin', None)
        if user:
            messages.success(request, f"Marked {user.first_name} {user.last_name} absent.")
        else:
            messages.success(request, "Marked attendee absent.")
    except Exception:
        pass

    return JsonResponse({'success': True, 'id': reg.id})