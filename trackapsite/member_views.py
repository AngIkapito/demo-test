from django.template.loader import render_to_string
from django.http import HttpResponse, FileResponse, Http404
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders
import io
import base64
from app.models import Membership, Member, MemberType, Organization, School_Year
from django.shortcuts import render,redirect, HttpResponse
from django.urls import path, include
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.urls import reverse
from app.models import CustomUser, Member, Membership, Member_Event_Registration, Event, Tags , Bulk_Event_Reg
from django.utils import timezone
from django.utils.safestring import mark_safe
#from django.http import JsonResponse
import json
from django.db import transaction
from django.conf import settings
import os
import pandas as pd
from openpyxl import load_workbook
from app.audit import audit_logger

@login_required(login_url='/')
def home(request):
    user = request.user
    membership_expiry = None
    membership_status = None
    member = None
    try:
        member = Member.objects.get(admin=user)
        membership = (
            Membership.objects.filter(member=member, status='Approved')
            .select_related('school_year')
            .order_by('-school_year__sy_end')
            .first()
        )
        if membership and membership.school_year:
            membership_expiry = membership.school_year.sy_end
            membership_status = membership.school_year.status
    except Member.DoesNotExist:
        pass
    except Exception:
        pass

    context = {
        'user': user,
        'membership_expiry': membership_expiry, 
        'membership_sy_status': membership_status,
    }
    # Compute total events attended by this member (approved and present)
    try:
        if member:
            attended_count = Member_Event_Registration.objects.filter(member_id=member, is_present=True, is_approved=True).count()
        else:
            attended_count = 0
    except Exception:
        attended_count = 0
    context['attended_count'] = attended_count
    # Build timeline entries from this member's Membership records
    timeline_entries = []
    try:
        if member:
            mem_qs = Membership.objects.filter(member=member).select_related('school_year').order_by('-school_year__sy_start')
            for m in mem_qs:
                sy = getattr(m, 'school_year', None)
                # Prefer membership.created_at as the timeline date
                created = getattr(m, 'created_at', None)
                date_str = None
                year = None
                if created:
                    try:
                        date_str = created.strftime('%b %d, %Y')
                        year = getattr(created, 'year', None)
                    except Exception:
                        date_str = str(created)
                        try:
                            year = int(str(created)[:4])
                        except Exception:
                            year = None
                else:
                    # fallback to school_year start year if created_at missing
                    if sy and getattr(sy, 'sy_start', None):
                        try:
                            year = getattr(sy.sy_start, 'year', None) or int(str(sy.sy_start)[:4])
                        except Exception:
                            year = None

                # Gather events the member joined for this school year where they were present
                joined_events = []
                try:
                    if member and sy:
                        regs = Member_Event_Registration.objects.filter(
                            member_id=member,
                            event__school_year=sy,
                            is_present=True,
                        ).select_related('event')
                        for r in regs:
                            ev = getattr(r, 'event', None)
                            joined_events.append({
                                'title': getattr(ev, 'title', '') if ev else '',
                                'date': getattr(ev, 'date', None).strftime('%b %d, %Y') if getattr(ev, 'date', None) else '',
                                'id': getattr(ev, 'id', None) if ev else None,
                            })
                except Exception:
                    joined_events = []

                timeline_entries.append({
                    'year': year,
                    'date': date_str,
                    'school_year_display': f"{getattr(sy, 'sy_start', '') and sy.sy_start.year}-{getattr(sy, 'sy_end', '') and sy.sy_end.year}" if sy and getattr(sy, 'sy_start', None) and getattr(sy, 'sy_end', None) else '',
                    'status': getattr(m, 'status', ''),
                    'id': getattr(m, 'id', None),
                    'notes': '',
                    'school_year_id': getattr(sy, 'id', None) if sy else None,
                    'joined_events': joined_events,
                })
    except Exception:
        timeline_entries = []

    context['timeline_entries'] = timeline_entries
    # Provide events JSON for the client-side calendar
    try:
        events_json = []
        for ev in Event.objects.all():
            date_iso = None
            if getattr(ev, 'date', None):
                try:
                    d = ev.date
                    if hasattr(d, 'date'):
                        date_iso = d.date().isoformat()
                    else:
                        date_iso = d.isoformat()
                except Exception:
                    date_iso = str(ev.date)

            banner_url = None
            try:
                b = getattr(ev, 'banner', None)
                if b:
                    banner_url = getattr(b, 'url', str(b))
            except Exception:
                banner_url = None

            events_json.append({
                'id': getattr(ev, 'id', None),
                'title': getattr(ev, 'title', '') or '',
                'date': date_iso,
                'banner_url': banner_url,
                'max_attendees': getattr(ev, 'max_attendees', None),
                'available_slots': getattr(ev, 'available_slots', None),
                'status': getattr(ev, 'status', None),
            })
    except Exception:
        events_json = []

    context['events_json'] = events_json
    return render(request,'member/home.html', context)

@login_required(login_url='/')
def generate_membership_certificate(request, membership_id):
    try:
        membership = Membership.objects.select_related('member', 'member__admin', 'member__organization', 'member__membershiptype', 'school_year').get(id=membership_id)
    except Membership.DoesNotExist:
        raise Http404("Membership not found")

    member = membership.member
    user = member.admin
    org = getattr(member, 'organization', None)
    school_year = membership.school_year
    membertype = getattr(member, 'membershiptype', None)

    # President name (static or from OfficerType/CustomUser if available)
    # Try to resolve the current president from CustomUser.user_type == 1
    president_name = "PSITE-CL President"
    try:
        prez = CustomUser.objects.filter(user_type=1).order_by('-id').first()
        if prez:
            president_name = f"{getattr(prez, 'first_name', '')} {getattr(prez, 'last_name', '')}".strip()
    except Exception:
        # keep default if lookup fails
        pass

    # Compose context for the certificate
    context = {
        'member_full_name': f"{user.first_name} {user.last_name}",
        'membership_type': membertype.name if membertype else '',
        'membership_id': membership.id,
        'school_name': org.name if org else '',
        'membership_start': school_year.sy_start.strftime('%b %d, %Y') if school_year and school_year.sy_start else '',
        'membership_end': school_year.sy_end.strftime('%b %d, %Y') if school_year and school_year.sy_end else '',
        'president_name': president_name,
        'issue_date': io.StringIO().getvalue() or '',  # fallback, will set below
        'tracking_number': f"MEM-{membership.id:06d}",
    }
    from datetime import datetime
    context['issue_date'] = datetime.now().strftime('%b %d, %Y')

    # Render HTML
    # Try to embed the logo as a base64 data URI so xhtml2pdf always sees it
    logo_data_uri = ''
    try:
        logo_rel = 'img/psitelogo.jpg'
        logo_path = finders.find(logo_rel)
        if not logo_path and getattr(settings, 'STATIC_ROOT', None):
            candidate = os.path.join(settings.STATIC_ROOT, logo_rel)
            if os.path.exists(candidate):
                logo_path = candidate
        if not logo_path:
            # try project static folder
            base_static = getattr(settings, 'BASE_DIR', '')
            candidate = os.path.join(base_static, 'static', logo_rel)
            if os.path.exists(candidate):
                logo_path = candidate

        if logo_path and os.path.exists(logo_path):
            with open(logo_path, 'rb') as lf:
                data = lf.read()
            encoded = base64.b64encode(data).decode('utf-8')
            logo_data_uri = f"data:image/jpeg;base64,{encoded}"
    except Exception:
        logo_data_uri = ''

    context['logo_data_uri'] = logo_data_uri
    # Embed footer/design image as base64 as well
    footer_design_data_uri = ''
    try:
        ft_rel = 'img/ftdesign.jpg'
        ft_path = finders.find(ft_rel)
        if not ft_path and getattr(settings, 'STATIC_ROOT', None):
            candidate = os.path.join(settings.STATIC_ROOT, ft_rel)
            if os.path.exists(candidate):
                ft_path = candidate
        if not ft_path:
            base_static = getattr(settings, 'BASE_DIR', '')
            candidate = os.path.join(base_static, 'static', ft_rel)
            if os.path.exists(candidate):
                ft_path = candidate

        if ft_path and os.path.exists(ft_path):
            with open(ft_path, 'rb') as ff:
                fdata = ff.read()
            fencoded = base64.b64encode(fdata).decode('utf-8')
            footer_design_data_uri = f"data:image/jpeg;base64,{fencoded}"
    except Exception:
        footer_design_data_uri = ''
    context['footer_design_uri'] = footer_design_data_uri
    html = render_to_string('certificate_template.html', context)

    # Generate PDF with a link_callback so xhtml2pdf can resolve static/media files
    result = io.BytesIO()

    def link_callback(uri, rel):
        """Convert HTML URIs to absolute system paths so xhtml2pdf can access them.

        Handles `MEDIA_URL` and `STATIC_URL` and falls back to the staticfiles finders.
        """
        # Media files
        if uri.startswith(getattr(settings, 'MEDIA_URL', '/media/')):
            path = os.path.join(getattr(settings, 'MEDIA_ROOT', ''), uri.replace(settings.MEDIA_URL, '').lstrip('/'))
            return path

        # Static files: try finders first
        static_path = finders.find(uri)
        if static_path:
            return static_path

        # If uri is prefixed with STATIC_URL, try to build path from STATIC_ROOT
        if uri.startswith(getattr(settings, 'STATIC_URL', '/static/')):
            path = os.path.join(getattr(settings, 'STATIC_ROOT', ''), uri.replace(settings.STATIC_URL, '').lstrip('/'))
            return path

        return uri

    pdf = pisa.pisaDocument(io.BytesIO(html.encode('utf-8')), result, link_callback=link_callback)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    # Save PDF to media/certificates
    cert_dir = os.path.join(settings.MEDIA_ROOT, 'certificates')
    os.makedirs(cert_dir, exist_ok=True)
    # Use tracking number as filename to match generated document label
    tracking = context.get('tracking_number', f"MEM-{membership.id:06d}")
    # sanitize filename (remove any unexpected slashes)
    safe_name = tracking.replace('/', '_')
    cert_filename = f"{safe_name}.pdf"
    cert_path = os.path.join(cert_dir, cert_filename)
    with open(cert_path, 'wb') as f:
        f.write(result.getvalue())

    # Save file path to membership.file_path
    rel_path = f"certificates/{cert_filename}"
    membership.file_path = rel_path
    membership.save(update_fields=['file_path'])

    # Return PDF as download
    return FileResponse(open(cert_path, 'rb'), as_attachment=True, filename=cert_filename)

@login_required(login_url='/')
def PROFILE(request):
    user = CustomUser.objects.get(id = request.user.id)
    
    context = {
        "user":user,

    }
    return render(request, 'member/profile.html', context)

@login_required(login_url='/')
def PROFILE_UPDATE(request):
    if request.method == "POST":
        profile_pic = request.FILES.get('profile_pic')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        change_password = request.POST.get('change_password')  # renamed field

        try:
            customuser = CustomUser.objects.get(id=request.user.id)

            customuser.first_name = first_name
            customuser.last_name = last_name
            customuser.email = email
            customuser.username = username

            # If change password field is filled, update password
            if change_password and change_password.strip() != "":
                customuser.set_password(change_password)
                customuser.save()
                update_session_auth_hash(request, customuser)  # keep user logged in
            else:
                customuser.save()

            if profile_pic and profile_pic != "":
                customuser.profile_pic = profile_pic
                customuser.save()

            messages.success(request, 'Your profile was updated successfully!')
            audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) updated their profile")
            return redirect('profile_update_member')  # redirect to the same update page
        except Exception as e:
            print("Error updating profile:", e)  # for debugging
            messages.error(request, 'Failed to update your profile')
            audit_logger.exception(f"Failed to update profile for user {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}): {e}")
            return redirect('profile_update_member')
    
    return render(request, 'member/profile.html') 


def basic_information(request):
    return render(request,'member/basic_information.html')




@login_required(login_url='/')
def MEMBER_EVENT_REG(request):
    """
    Handles the Member Event Registration form:
    - Displays active event details automatically.
    - Gets the logged-in user's ID.
    - Saves a registration to the Member_Event_Registration table.
    - Updates available_slots instead of modifying max_attendees.
    - Prevents duplicate registrations by the same user.
    - Shows a message when no active event exists.
    """
    # ✅ Fetch the active event (latest if multiple)
    event = Event.objects.filter(status='active').order_by('-date').first()

    # ✅ If no active event found
    if not event:
        messages.error(request, "There are currently no active events available for registration.")
        audit_logger.warning(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) attempted member event registration but no active event exists")
        return render(request, 'officer/member_event_reg.html', {'event': None})

    if request.method == 'POST':
        # map the current logged-in user to the Member record
        try:
            member = Member.objects.get(admin=request.user)
        except Member.DoesNotExist:
            messages.error(request, "Member profile not found for this user.")
            return redirect('member_event_reg_member')

        # ✅ Check for duplicate registration using member_id FK
        already_registered = Member_Event_Registration.objects.filter(
            member_id=member,
            event_id=event.id,
            status='registered'
        ).exists()

        if already_registered:
            messages.warning(request, "You are already registered for this event.")
            audit_logger.warning(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) attempted duplicate registration for event id={getattr(event, 'id', None)}")
            return redirect('member_event_reg_member')

        # ✅ Determine available slots (default to max_attendees if available_slots not set)
        available_slots = getattr(event, 'available_slots', event.max_attendees)

        if available_slots <= 0:
            # set both `is_closed` and `is_full` when no slots remain
            if getattr(event, 'is_closed', 0) != 1 or not getattr(event, 'is_full', False):
                try:
                    event.is_closed = 1
                    event.is_full = True
                    event.save(update_fields=['is_closed', 'is_full'])
                except Exception:
                    # silent fallback if fields don't exist
                    try:
                        event.is_closed = 1
                        event.save(update_fields=['is_closed'])
                    except Exception:
                        pass
            messages.error(request, "Registration closed. The event is full.")
            return redirect('member_event_reg_member')

        try:
            # ✅ Create the registration record
            Member_Event_Registration.objects.create(
                member_id=member,
                event_id=event.id,
                date_created=timezone.now(),
                status='registered'
            )

            # ✅ Recalculate available slots: max_attendees - total registered
            total_registered = Member_Event_Registration.objects.filter(
                event_id=event.id, status='registered'
            ).count()
            new_available_slots = event.max_attendees - total_registered
            # Prevent negative available_slots which would cause DB errors
            if new_available_slots < 0:
                new_available_slots = 0

            # ✅ Update event.available_slots and is_closed flag (do NOT change status)
            event.available_slots = new_available_slots
            # set is_closed and is_full flags when no slots remain
            if new_available_slots <= 0:
                try:
                    event.is_closed = 1
                except Exception:
                    pass
                try:
                    event.is_full = True
                except Exception:
                    pass
            else:
                try:
                    event.is_closed = 0
                except Exception:
                    pass
                try:
                    event.is_full = False
                except Exception:
                    pass
            # try to save all fields; fall back to smaller sets
            try:
                event.save(update_fields=['available_slots', 'is_closed', 'is_full'])
            except Exception:
                try:
                    event.save(update_fields=['available_slots', 'is_closed'])
                except Exception:
                    event.save(update_fields=['available_slots'])

            messages.success(request, "Registration successful!")
            audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) registered for event id={getattr(event, 'id', None)} title={getattr(event, 'title', None)} member_id={getattr(member, 'id', None)}")
            return redirect('member_event_reg_member')

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            audit_logger.exception(f"Error registering user {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) for event id={getattr(event, 'id', None)}: {e}")
            return redirect('member_event_reg_member')
    context = {'event': event}
    return render(request, 'member/member_event_reg.html', context)




def BULK_EVENT_REG(request):
    """Render the bulk event registration page."""
    # Fetch the currently active event (latest by date if multiple)
    active_event = Event.objects.filter(status='active').order_by('-date').first()
    return render(request, 'member/bulk_event_reg.html', {
        'event': active_event
    })

@login_required(login_url='/')

def UPLOAD_BULK_EVENT_REG(request):
    """Render the bulk event registration page and handle file uploads.

    POST behavior (from the template buttons):
    - action=upload : save uploaded file to `MEDIA_ROOT/bulk_event/bulk_event_registration{ext}`
    - action=view   : same as upload (keeps behavior simple); view logic can be expanded later
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        # Only perform file saving when action is explicitly 'upload'
        if action != 'upload':
            messages.info(request, 'No upload performed.')
            return redirect('upload_bulk_event_reg_member')

        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            messages.warning(request, 'No file selected to upload.')
            return redirect('upload_bulk_event_reg_member')

        # ensure media bulk_event directory exists
        media_dir = os.path.join(getattr(settings, 'MEDIA_ROOT', 'media'), 'bulk_event')
        try:
            os.makedirs(media_dir, exist_ok=True)
        except Exception:
            messages.error(request, 'Failed to create media directory.')
            return redirect('upload_bulk_event_reg_member')

        # save using the original uploaded filename (basename only)
        original_name = os.path.basename(uploaded.name)
        if not original_name:
            # fallback name
            original_name = 'uploaded_file.csv'
        filename = original_name
        dest_path = os.path.join(media_dir, filename)

        try:
            with open(dest_path, 'wb+') as dest:
                for chunk in uploaded.chunks():
                    dest.write(chunk)
        except Exception as e:
            messages.error(request, f'Failed to save uploaded file: {e}')
            return redirect('upload_bulk_event_reg_member')

        # Parse the uploaded file with pandas to build preview_rows
        preview_rows = []
        try:
            ext = os.path.splitext(dest_path)[1].lower()
            wanted = [
                'last_name','first_name','middle','contact_number','email',
                'attending_as','is_competitor','if_competitor','is_coach','if_coach','tshirt_size'
            ]

            if ext == '.csv':
                df = pd.read_csv(dest_path, header=14)
                df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
                # Fix common typos for 'if_competitor'
                typo_variants = ['if_compeitor','if_comptetitor','if_competior','if _compeitor']
                for tv in typo_variants:
                    if tv in df.columns:
                        df = df.rename(columns={tv: 'if_competitor'})

                cols = [c for c in wanted if c in df.columns]
                df = df[cols]
                for _, row in df.head(1000).iterrows():
                    item = {k: ('' if pd.isna(row.get(k)) else str(row.get(k))) for k in wanted}
                    # only include rows with a non-empty last_name
                    if str(item.get('last_name', '')).strip():
                        preview_rows.append(item)
            else:
                # Use openpyxl to read Excel with data_only=True
                wb = load_workbook(dest_path, data_only=True)
                ws = wb.active
                # Header row is 15 (1-based); collect headers and normalize
                header_row_index = 15
                headers = []
                for cell in ws[header_row_index]:
                    val = '' if cell.value is None else str(cell.value)
                    headers.append(val.strip().lower().replace(' ', '_'))

                # Fix common typos in headers list
                headers = [
                    ('if_competitor' if h in ['if_compeitor','if_comptetitor','if_competior','if _compeitor'] else h)
                    for h in headers
                ]

                # Build mapping col_index -> header
                header_map = {idx: h for idx, h in enumerate(headers)}

                # Iterate rows after header (from 16 to max_row)
                max_rows = min(ws.max_row, 14 + 1000)  # limit ~1000 data rows
                for r in range(16, max_rows + 1):
                    row_vals = {}
                    for idx, cell in enumerate(ws[r]):
                        h = header_map.get(idx)
                        if h:
                            val = cell.value
                            row_vals[h] = '' if val is None else str(val)

                    item = {k: row_vals.get(k, '') for k in wanted}
                    # only include rows with a non-empty last_name
                    if str(item.get('last_name', '')).strip():
                        preview_rows.append(item)
        except Exception as e:
            messages.error(request, f'Failed to parse file for preview: {e}')
            preview_rows = []

        messages.success(request, f'File uploaded: {filename}. Parsed {len(preview_rows)} rows with last_name for preview.')
        audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) uploaded bulk event file {filename} parsed_rows={len(preview_rows)}")
        return render(request, 'officer/bulk_event_reg.html', {
            'preview_rows': preview_rows,
            'event': Event.objects.filter(status='active').order_by('-date').first()
        })

    return render(request, 'officer/bulk_event_reg.html')

def SAVE_BULK_EVENT_REG(request):
    """Persist uploaded preview rows into Bulk_Event_Reg.

    Expects POST with 'preview_json' containing the preview rows shown in the table.
    Only rows with non-empty last_name should be processed (already filtered client-side).
    """
    if request.method != 'POST':
        return redirect('bulk_event_reg_member')

    preview_json = request.POST.get('preview_json')
    if not preview_json:
        messages.error(request, 'No preview data to save.')
        return redirect('bulk_event_reg_member')

    try:
        # The template uses json_script; adapt to plain JSON if provided
        # If it contains a <script> tag, extract inner text; otherwise treat as JSON
        data_str = preview_json
        # Attempt to load JSON
        rows = json.loads(data_str)
        if not isinstance(rows, list):
            messages.error(request, 'Invalid preview data format.')
            return redirect('bulk_event_reg_member')
    except Exception as e:
        messages.error(request, f'Failed to parse preview data: {e}')
        return redirect('bulk_event_reg_member')

    # Fetch active event to associate registrations
    event = Event.objects.filter(status='active').order_by('-date').first()
    if not event:
        messages.error(request, 'No active event to associate registrations.')
        return redirect('bulk_event_reg_member')

    saved = 0
    skipped = 0

    # Persist within a transaction for reliability
    try:
        with transaction.atomic():
            def as_bool(val):
                if val is None:
                    return False
                s = str(val).strip().lower()
                return s in ['yes','true','1','y','t']

            # Resolve Member.id for the current user (instead of CustomUser.id)
            member_id = (
                Member.objects.filter(admin_id=getattr(request.user, 'id', None))
                .values_list('id', flat=True)
                .first()
            )

            for r in rows:
                try:
                    # Only save fields that exist; map expected keys
                    reg = Bulk_Event_Reg.objects.create(
                        event_id=event.id,
                        # Store Member.id in registered_by (FK to Member)
                        registered_by_id=member_id,
                        last_name=r.get('last_name', '') or '',
                        first_name=r.get('first_name', '') or '',
                        middle=r.get('middle', '') or '',
                        contact_number=r.get('contact_number', '') or '',
                        email=r.get('email', '') or '',
                        attending_as=r.get('attending_as', '') or '',
                        is_competitor=as_bool(r.get('is_competitor', False)),
                        if_competitor=r.get('if_competitor', '') or '',
                        is_coach=as_bool(r.get('is_coach', False)),
                        if_coach=r.get('if_coach', '') or '',
                        tshirt_size=r.get('tshirt_size', '') or '',
                    )
                    saved += 1
                except Exception:
                    skipped += 1
                    continue
    except Exception as e:
        messages.error(request, f'Failed to save registrations: {e}')
        return redirect('bulk_event_reg_member')

    messages.success(request, f'Saved {saved} registrations. Skipped {skipped}.')
    audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) saved bulk registrations for event id={getattr(event, 'id', None)} saved={saved} skipped={skipped}")
    return redirect('bulk_event_reg_member')

