from django.shortcuts import render,redirect, HttpResponse, get_object_or_404
from django.urls import path, include, reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from app.models import CustomUser, Event, School_Year,Announcement, Salutation,Organization, MemberType, MembershipType, Member, OfficerType, Region, Membership, Member_Event_Registration, Bulk_Event_Reg, Tags
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.http import JsonResponse
import datetime
from django.views.decorators.http import require_http_methods
from django.core.mail import send_mail
from django.utils.crypto import get_random_string
import json
from django.db import transaction
from django.conf import settings
import os
import qrcode
import pandas as pd
from openpyxl import load_workbook
from app.audit import audit_logger
from django.template.loader import render_to_string
from django.http import FileResponse, Http404
from django.contrib.staticfiles import finders
from xhtml2pdf import pisa
import io
import base64

@login_required(login_url='/')
def home(request):
    # Get the logged-in user
    user = request.user
    membership_expiry = None
    membership_status = None
    # Find the Member object for this user
    try:
        member = Member.objects.get(admin=user)
        # Get the latest approved membership for this member
        membership = (
            Membership.objects.filter(member=member, status='Approved')
            .select_related('school_year')
            .order_by('-school_year__sy_end')
            .first()
        )
        if membership and membership.school_year:
            membership_expiry = membership.school_year.sy_end
            membership_status = membership.school_year.status
    except Member.DoesNotExist:
        pass
    except Exception:
        pass

    context = {
        'user': user,
        'membership_expiry': membership_expiry,
        'membership_sy_status': membership_status,
    }
    # Build timeline entries from this member's Membership records (same behavior as member_views.home)
    timeline_entries = []
    try:
        if 'member' in locals() and member:
            mem_qs = Membership.objects.filter(member=member).select_related('school_year').order_by('-school_year__sy_start')
            for m in mem_qs:
                sy = getattr(m, 'school_year', None)
                created = getattr(m, 'created_at', None)
                date_str = None
                year = None
                if created:
                    try:
                        date_str = created.strftime('%b %d, %Y')
                        year = getattr(created, 'year', None)
                    except Exception:
                        date_str = str(created)
                        try:
                            year = int(str(created)[:4])
                        except Exception:
                            year = None
                else:
                    if sy and getattr(sy, 'sy_start', None):
                        try:
                            year = getattr(sy.sy_start, 'year', None) or int(str(sy.sy_start)[:4])
                        except Exception:
                            year = None

                joined_events = []
                try:
                    if member and sy:
                        regs = Member_Event_Registration.objects.filter(
                            member_id=member,
                            event__school_year=sy,
                            is_present=True,
                        ).select_related('event')
                        for r in regs:
                            ev = getattr(r, 'event', None)
                            joined_events.append({
                                'title': getattr(ev, 'title', '') if ev else '',
                                'date': getattr(ev, 'date', None).strftime('%b %d, %Y') if getattr(ev, 'date', None) else '',
                                'id': getattr(ev, 'id', None) if ev else None,
                            })
                except Exception:
                    joined_events = []

                timeline_entries.append({
                    'year': year,
                    'date': date_str,
                    'school_year_display': f"{getattr(sy, 'sy_start', '') and sy.sy_start.year}-{getattr(sy, 'sy_end', '') and sy.sy_end.year}" if sy and getattr(sy, 'sy_start', None) and getattr(sy, 'sy_end', None) else '',
                    'status': getattr(m, 'status', ''),
                    'id': getattr(m, 'id', None),
                    'notes': '',
                    'school_year_id': getattr(sy, 'id', None) if sy else None,
                    'joined_events': joined_events,
                })
    except Exception:
        timeline_entries = []

    context['timeline_entries'] = timeline_entries
    return render(request, 'officer/home.html', context)


@login_required(login_url='/')
def generate_membership_certificate(request, membership_id):
    try:
        membership = Membership.objects.select_related('member', 'member__admin', 'member__organization', 'member__membershiptype', 'school_year').get(id=membership_id)
    except Membership.DoesNotExist:
        raise Http404("Membership not found")

    member = membership.member
    user = member.admin
    org = getattr(member, 'organization', None)
    school_year = membership.school_year
    membertype = getattr(member, 'membershiptype', None)

    president_name = "PSITE-CL President"
    try:
        prez = CustomUser.objects.filter(user_type=1).order_by('-id').first()
        if prez:
            president_name = f"{getattr(prez, 'first_name', '')} {getattr(prez, 'last_name', '')}".strip()
    except Exception:
        pass

    context = {
        'member_full_name': f"{user.first_name} {user.last_name}",
        'membership_type': membertype.name if membertype else '',
        'membership_id': membership.id,
        'school_name': org.name if org else '',
        'membership_start': school_year.sy_start.strftime('%b %d, %Y') if school_year and school_year.sy_start else '',
        'membership_end': school_year.sy_end.strftime('%b %d, %Y') if school_year and school_year.sy_end else '',
        'president_name': president_name,
        'issue_date': io.StringIO().getvalue() or '',
        'tracking_number': f"MEM-{membership.id:06d}",
    }
    from datetime import datetime
    context['issue_date'] = datetime.now().strftime('%b %d, %Y')

    logo_data_uri = ''
    try:
        logo_rel = 'img/psitelogo.jpg'
        logo_path = finders.find(logo_rel)
        if not logo_path and getattr(settings, 'STATIC_ROOT', None):
            candidate = os.path.join(settings.STATIC_ROOT, logo_rel)
            if os.path.exists(candidate):
                logo_path = candidate
        if not logo_path:
            base_static = getattr(settings, 'BASE_DIR', '')
            candidate = os.path.join(base_static, 'static', logo_rel)
            if os.path.exists(candidate):
                logo_path = candidate

        if logo_path and os.path.exists(logo_path):
            with open(logo_path, 'rb') as lf:
                data = lf.read()
            encoded = base64.b64encode(data).decode('utf-8')
            logo_data_uri = f"data:image/jpeg;base64,{encoded}"
    except Exception:
        logo_data_uri = ''

    context['logo_data_uri'] = logo_data_uri

    footer_design_data_uri = ''
    try:
        ft_rel = 'img/ftdesign.jpg'
        ft_path = finders.find(ft_rel)
        if not ft_path and getattr(settings, 'STATIC_ROOT', None):
            candidate = os.path.join(settings.STATIC_ROOT, ft_rel)
            if os.path.exists(candidate):
                ft_path = candidate
        if not ft_path:
            base_static = getattr(settings, 'BASE_DIR', '')
            candidate = os.path.join(base_static, 'static', ft_rel)
            if os.path.exists(candidate):
                ft_path = candidate

        if ft_path and os.path.exists(ft_path):
            with open(ft_path, 'rb') as ff:
                fdata = ff.read()
            fencoded = base64.b64encode(fdata).decode('utf-8')
            footer_design_data_uri = f"data:image/jpeg;base64,{fencoded}"
    except Exception:
        footer_design_data_uri = ''
    context['footer_design_uri'] = footer_design_data_uri

    html = render_to_string('certificate_template.html', context)

    result = io.BytesIO()

    def link_callback(uri, rel):
        if uri.startswith(getattr(settings, 'MEDIA_URL', '/media/')):
            path = os.path.join(getattr(settings, 'MEDIA_ROOT', ''), uri.replace(settings.MEDIA_URL, '').lstrip('/'))
            return path
        static_path = finders.find(uri)
        if static_path:
            return static_path
        if uri.startswith(getattr(settings, 'STATIC_URL', '/static/')):
            path = os.path.join(getattr(settings, 'STATIC_ROOT', ''), uri.replace(settings.STATIC_URL, '').lstrip('/'))
            return path
        return uri

    pdf = pisa.pisaDocument(io.BytesIO(html.encode('utf-8')), result, link_callback=link_callback)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)

    cert_dir = os.path.join(settings.MEDIA_ROOT, 'certificates')
    os.makedirs(cert_dir, exist_ok=True)
    tracking = context.get('tracking_number', f"MEM-{membership.id:06d}")
    safe_name = tracking.replace('/', '_')
    cert_filename = f"{safe_name}.pdf"
    cert_path = os.path.join(cert_dir, cert_filename)
    with open(cert_path, 'wb') as f:
        f.write(result.getvalue())

    rel_path = f"certificates/{cert_filename}"
    membership.file_path = rel_path
    membership.save(update_fields=['file_path'])

    return FileResponse(open(cert_path, 'rb'), as_attachment=True, filename=cert_filename)


@login_required(login_url='/')
def PROFILE(request):
    user = CustomUser.objects.get(id = request.user.id)
    
    context = {
        "user":user,

    }
    return render(request, 'officer/profile.html', context)


@login_required(login_url='/')
def PROFILE_UPDATE(request):
    if request.method == "POST":
        profile_pic = request.FILES.get('profile_pic')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        username = request.POST.get('username')
        change_password = request.POST.get('change_password')  # renamed field

        try:
            customuser = CustomUser.objects.get(id=request.user.id)

            customuser.first_name = first_name
            customuser.last_name = last_name
            customuser.email = email
            customuser.username = username

            # If change password field is filled, update password
            if change_password and change_password.strip() != "":
                customuser.set_password(change_password)
                customuser.save()
                update_session_auth_hash(request, customuser)  # keep user logged in
            else:
                customuser.save()

            if profile_pic and profile_pic != "":
                customuser.profile_pic = profile_pic
                customuser.save()

            messages.success(request, 'Your profile was updated successfully!')
            audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) updated their profile (officer)")
            return redirect('profile_update_officer')  # redirect to the same update page
        except Exception as e:
            print("Error updating profile:", e)  # for debugging
            messages.error(request, 'Failed to update your profile')
            audit_logger.exception(f"Failed to update profile for user {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) (officer): {e}")
            return redirect('profile_update_officer')
    
    return render(request, 'officer/profile.html') 


def VIEWALL_EVENT(request):
    # Fetch all events with related school year to reduce queries
    events = Event.objects.select_related('school_year').all()
    return render(request, 'officer/viewall_event.html', {'events': events})




@login_required(login_url='/')
def ADD_EVENT(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        theme = request.POST.get('theme')
        date = request.POST.get('date')
        location = request.POST.get('location')
        max_attendees = request.POST.get('max_attendees')
        registration_fee = request.POST.get('registration_fee')
        chair_id = request.POST.get('chair')
        co_chair_id = request.POST.get('co_chair')
        evaluation_link = request.POST.get('evaluation_link')

        banner = request.FILES.get('banner')
        tag_id = request.POST.get('tag')

        try:
            active_schoolyear = School_Year.objects.get(status=1)
        except School_Year.DoesNotExist:
            messages.error(request, "No active school year found. Please activate a school year first.")
            return redirect('add_event')

        Event.objects.filter(status='active').update(status='inactive')

        try:
            max_attendees_val = int(max_attendees) if max_attendees not in (None, '') else 0
        except Exception:
            max_attendees_val = 0
        try:
            registration_fee_val = float(registration_fee) if registration_fee not in (None, '') else 0.0
        except Exception:
            registration_fee_val = 0.0

        event = Event(
            title=title,
            theme=theme,
            date=date,
            location=location,
            max_attendees=max_attendees_val,
            registration_fee=registration_fee_val,
            chair_id=chair_id,
            co_chair_id=co_chair_id,
            evaluation_link=evaluation_link,
            banner=banner,
            created_by_id=request.user.id,
            created_at=timezone.now(),
            updated_at=timezone.now(),
            school_year=active_schoolyear,
            status='active',
            available_slots=max_attendees_val,
            tags_id=tag_id if tag_id not in (None, '') else None
        )
        event.save()

        try:
            reg_path = reverse('registration_event') + f'?event_id={event.id}'
            registration_url = request.build_absolute_uri(reg_path)
            event.registration_link = registration_url
            event.save()
        except Exception as e:
            print('Failed to set registration_link for event', event.id, e)

        try:
            qr_data = event.registration_link if event.registration_link else f"EVENT:{event.id}"
            img = qrcode.make(qr_data)
            qr_dir = os.path.join(settings.MEDIA_ROOT, 'qr_codes')
            os.makedirs(qr_dir, exist_ok=True)
            filename = f"event_{event.id}_qr.png"
            dest_path = os.path.join(qr_dir, filename)
            img.save(dest_path)
            event.qr_code = os.path.join('qr_codes', filename).replace('\\', '/')
            event.save()
        except Exception as e:
            print('Failed to generate QR code for event', event.id, e)

        bulk_template = request.FILES.get('bulk_template')
        if bulk_template:
            try:
                dest_dir = os.path.join(settings.MEDIA_ROOT, 'bulk_template')
                os.makedirs(dest_dir, exist_ok=True)
                timestamp = timezone.now().strftime('%Y%m%d%H%M%S')
                filename = f"{timestamp}_{bulk_template.name}"
                dest_path = os.path.join(dest_dir, filename)
                with open(dest_path, 'wb+') as dst:
                    for chunk in bulk_template.chunks():
                        dst.write(chunk)
                event.template_path = os.path.join('bulk_template', filename).replace('\\', '/')
                event.save()
            except Exception as e:
                print('Failed to save bulk_template:', e)

        messages.success(request, f'Event added successfully for cycle {active_schoolyear.sy_start.year} - {active_schoolyear.sy_end.year}!')
        audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) added event id={event.id} title={event.title}")
        return redirect('viewall_event')

    members = Member.objects.all()
    officertypes = OfficerType.objects.all()
    custom_users = CustomUser.objects.filter(user_type__in=[1, 2], is_superuser=0)
    tags = Tags.objects.all()

    for user in custom_users:
        try:
            member = members.get(admin_id=user.id)
            officer_type = officertypes.get(id=member.officertype_id)
            user.officertype_name = officer_type.name
        except (Member.DoesNotExist, OfficerType.DoesNotExist):
            user.officertype_name = "N/A"

    active_schoolyear = School_Year.objects.filter(status=1).first()

    return render(request, 'officer/add_event.html', {
        'members': members,
        'custom_users': custom_users,
        'officertypes': officertypes,
        'active_schoolyear': active_schoolyear,
        'tags': tags
    })
def MEMBER_EVENT_REG(request):
    """
    Handles the Member Event Registration form:
    - Displays active event details automatically.
    - Gets the logged-in user's ID.
    - Saves a registration to the Member_Event_Registration table.
    - Updates available_slots instead of modifying max_attendees.
    - Prevents duplicate registrations by the same user.
    - Shows a message when no active event exists.
    """
    # ✅ Fetch the active event (latest if multiple)
    event = Event.objects.filter(status='active').order_by('-date').first()

    # ✅ If no active event found
    if not event:
        messages.error(request, "There are currently no active events available for registration.")
        audit_logger.warning(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) attempted officer event registration but no active event exists")
        return render(request, 'officer/member_event_reg.html', {'event': None})

    if request.method == 'POST':
        # Resolve Member record for the logged-in user
        member_obj = Member.objects.filter(admin_id=getattr(request.user, 'id', None)).first()
        if not member_obj:
            messages.error(request, "No member profile found for the current user.")
            return redirect('member_event_reg_officer')

        # Check for duplicate registration using Member FK
        already_registered = Member_Event_Registration.objects.filter(
            member_id=member_obj,
            event=event,
            status='registered'
        ).exists()

        if already_registered:
            messages.warning(request, "You are already registered for this event.")
            audit_logger.warning(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) attempted duplicate officer registration for event id={getattr(event, 'id', None)}")
            return redirect('member_event_reg_officer')

        try:
            # Create the registration record referencing Member (member_id)
            Member_Event_Registration.objects.create(
                member_id=member_obj,
                event=event,
                date_created=timezone.now(),
                status='registered'
            )

            messages.success(request, "Registration successful!")
            audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) registered (officer) for event id={getattr(event, 'id', None)} title={getattr(event, 'title', None)} member_id={getattr(member_obj, 'id', None)}")
            return redirect('member_event_reg_officer')

        except Exception as e:
            messages.error(request, f"An error occurred: {e}")
            audit_logger.exception(f"Error registering officer user {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) for event id={getattr(event, 'id', None)}: {e}")
            return redirect('member_event_reg_officer')
    context = {'event': event}
    return render(request, 'officer/member_event_reg.html', context)


@login_required(login_url='/')
@require_http_methods(["GET", "POST"])
def MEMBERSHIP_APPROVAL(request):
    """
    View for Membership Registration Approval.
    GET: Fetch all members and map membership status.
    POST: Update membership status using membership.member_id for uniqueness.
    Sends emails using CustomUser email linked via Member.admin.
    """
    if request.method == "POST":
        # Use membership.member_id as unique identifier
        member_id = request.POST.get("member_id")
        action = request.POST.get("action")  # "approve" or "decline"

        # Find Member
        member = get_object_or_404(Member, id=member_id)
        # Get CustomUser details for name/email
        user = member.admin

        # Get related Membership
        # A member may have multiple Membership records (new + renew). Select the most
        # recent one to act upon. Prefer pending records if present; otherwise fall back
        # to the latest by id.
        membership_qs = Membership.objects.filter(member_id=member.id).order_by('-id')
        # Prefer a PENDING membership if one exists
        membership = membership_qs.filter(status__iexact='PENDING').first()
        if not membership:
            # Fall back to the latest membership record for this member
            membership = membership_qs.first()

        if not membership:
            messages.error(request, f"No membership record found for {user.first_name} {user.last_name}.")
            return redirect("membership_approval")

        if action == "approve":
            # Approve the membership record
            membership.status = "APPROVED"
            # record who processed this approval (store the id directly)
            try:
                membership.processed_by_id = getattr(request.user, 'id', None)
            except Exception:
                membership.processed_by_id = None
            membership.save()

            # Activate the user account
            user.is_active = 1
            user.save()

            # If this is a NEW membership (membertype != 2) generate credentials and email them.
            # For RENEWALS (membertype_id == 2) do NOT create a new password — just notify the user.
            if getattr(membership, 'membertype_id', None) != 2:
                # ✅ Generate new password for NEW registrations
                password = get_random_string(length=10)
                user.set_password(password)
                user.save()

                # ✅ Send approval email with username and password
                send_mail(
                    subject="Membership Approved",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"Your membership has been approved.\n"
                        f"Here are your login credentials:\n\n"
                        f"Username: {user.username}\n"
                        f"Password: {password}\n\n"
                        f"Please login and change your password as soon as possible."
                    ),
                    from_email="yourgmail@gmail.com",  # Replace with your EMAIL_HOST_USER
                    recipient_list=[user.email],
                    fail_silently=False,
                )

                messages.success(
                    request,
                    f"Membership for {user.first_name} {user.last_name} approved, user activated, and email with credentials sent."
                )
            else:
                # Renewal: do not generate a new password. Send a simple approval notification.
                send_mail(
                    subject="Membership Renewal Approved",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"Your membership renewal has been approved.\n"
                        f"No changes were made to your login credentials.\n\n"
                        f"Thank you."
                    ),
                    from_email="yourgmail@gmail.com",  # Replace with your EMAIL_HOST_USER
                    recipient_list=[user.email],
                    fail_silently=False,
                )

                messages.success(
                    request,
                    f"Membership renewal for {user.first_name} {user.last_name} approved; user retains existing credentials."
                )

        elif action == "decline":
            membership.status = "DECLINED"
            # record who processed this decline (store the id directly)
            try:
                membership.processed_by_id = getattr(request.user, 'id', None)
            except Exception:
                membership.processed_by_id = None
            membership.save()

            # Send decline email notification (logic adapted from hoo_views.py)
            try:
                send_mail(
                    subject="Membership Declined",
                    message=(
                        f"Dear {user.first_name} {user.last_name},\n\n"
                        f"We regret to inform you that your membership application has been declined.\n\n"
                        f"If you have questions or believe this is a mistake, please contact the organization.\n\n"
                        f"Thank you."
                    ),
                    from_email="yourgmail@gmail.com",  # Replace with your EMAIL_HOST_USER
                    recipient_list=[user.email],
                    fail_silently=False,
                )
            except Exception as e:
                audit_logger.exception(f"Failed to send decline email to {user.email}: {e}")

            processor_id = getattr(request.user, 'id', 'unknown')
            messages.warning(
                request,
                f"Membership for {user.first_name} {user.last_name} declined. Processed by id: {processor_id} and email sent."
            )
        else:
            messages.error(request, "Invalid action.")

        return redirect("membership_approval")

    # --- GET logic ---
    # Only consider memberships that are either NEW (membertype_id=1) or RENEW (membertype_id=2)
    memberships = (
        Membership.objects.filter(membertype_id__in=[1, 2])
        .order_by('-id')
        .only('member_id', 'status', 'proof_of_payment', 'membertype_id')
    )

    # Build maps using the latest membership per member (ordered by -id above)
    membership_status_map = {}
    membership_payment_map = {}
    membership_type_map = {}
    for m in memberships:
        if m.member_id not in membership_status_map:
            membership_status_map[m.member_id] = m.status
            membership_payment_map[m.member_id] = m.proof_of_payment
            membership_type_map[m.member_id] = getattr(m, 'membertype_id', None)

    # Get only members who have a relevant membership (new or renew)
    member_ids = list(membership_status_map.keys())
    members = Member.objects.select_related('admin', 'organization', 'membershiptype').filter(id__in=member_ids)

    # Add dynamic attributes for display using the maps
    for member in members:
        member.status = membership_status_map.get(member.id, 'PENDING')
        member.proof_of_payment = membership_payment_map.get(member.id, None)
        member.latest_membertype_id = membership_type_map.get(member.id)

    context = {
        'schoolyear': School_Year.objects.all(),
        'users': CustomUser.objects.all(),
        'members': members,
        'membership_status_map': membership_status_map,
        'memberships': memberships,
        'membership_types': MembershipType.objects.all(),
        'member_types': MemberType.objects.all(),
        'organization': Organization.objects.all(),
    }

    return render(request, 'officer/membership_approval.html', context)


def BULK_EVENT_REG(request):
    """Render the bulk event registration page."""
    # Fetch the currently active event (latest by date if multiple)
    active_event = Event.objects.filter(status='active').order_by('-date').first()
    return render(request, 'officer/bulk_event_reg.html', {
        'event': active_event
    })

@login_required(login_url='/')

def UPLOAD_BULK_EVENT_REG(request):
    """Render the bulk event registration page and handle file uploads.

    POST behavior (from the template buttons):
    - action=upload : save uploaded file to `MEDIA_ROOT/bulk_event/bulk_event_registration{ext}`
    - action=view   : same as upload (keeps behavior simple); view logic can be expanded later
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        # Only perform file saving when action is explicitly 'upload'
        if action != 'upload':
            messages.info(request, 'No upload performed.')
            return redirect('upload_bulk_event_reg_officer')

        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            messages.warning(request, 'No file selected to upload.')
            return redirect('upload_bulk_event_reg_officer')

        # ensure media bulk_event directory exists
        media_dir = os.path.join(getattr(settings, 'MEDIA_ROOT', 'media'), 'bulk_event')
        try:
            os.makedirs(media_dir, exist_ok=True)
        except Exception:
            messages.error(request, 'Failed to create media directory.')
            return redirect('upload_bulk_event_reg_officer')

        # save using the original uploaded filename (basename only)
        original_name = os.path.basename(uploaded.name)
        if not original_name:
            # fallback name
            original_name = 'uploaded_file.csv'
        filename = original_name
        dest_path = os.path.join(media_dir, filename)

        try:
            with open(dest_path, 'wb+') as dest:
                for chunk in uploaded.chunks():
                    dest.write(chunk)
        except Exception as e:
            messages.error(request, f'Failed to save uploaded file: {e}')
            return redirect('upload_bulk_event_reg_officer')

        # Parse the uploaded file with pandas to build preview_rows
        preview_rows = []
        try:
            ext = os.path.splitext(dest_path)[1].lower()
            wanted = [
                'last_name','first_name','middle','contact_number','email',
                'attending_as','is_competitor','if_competitor','is_coach','if_coach','tshirt_size'
            ]

            if ext == '.csv':
                df = pd.read_csv(dest_path, header=14)
                df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
                # Fix common typos for 'if_competitor'
                typo_variants = ['if_compeitor','if_comptetitor','if_competior','if _compeitor']
                for tv in typo_variants:
                    if tv in df.columns:
                        df = df.rename(columns={tv: 'if_competitor'})

                cols = [c for c in wanted if c in df.columns]
                df = df[cols]
                for _, row in df.head(1000).iterrows():
                    item = {k: ('' if pd.isna(row.get(k)) else str(row.get(k))) for k in wanted}
                    # only include rows with a non-empty last_name
                    if str(item.get('last_name', '')).strip():
                        preview_rows.append(item)
            else:
                # Use openpyxl to read Excel with data_only=True
                wb = load_workbook(dest_path, data_only=True)
                ws = wb.active
                # Header row is 15 (1-based); collect headers and normalize
                header_row_index = 15
                headers = []
                for cell in ws[header_row_index]:
                    val = '' if cell.value is None else str(cell.value)
                    headers.append(val.strip().lower().replace(' ', '_'))

                # Fix common typos in headers list
                headers = [
                    ('if_competitor' if h in ['if_compeitor','if_comptetitor','if_competior','if _compeitor'] else h)
                    for h in headers
                ]

                # Build mapping col_index -> header
                header_map = {idx: h for idx, h in enumerate(headers)}

                # Iterate rows after header (from 16 to max_row)
                max_rows = min(ws.max_row, 14 + 1000)  # limit ~1000 data rows
                for r in range(16, max_rows + 1):
                    row_vals = {}
                    for idx, cell in enumerate(ws[r]):
                        h = header_map.get(idx)
                        if h:
                            val = cell.value
                            row_vals[h] = '' if val is None else str(val)

                    item = {k: row_vals.get(k, '') for k in wanted}
                    # only include rows with a non-empty last_name
                    if str(item.get('last_name', '')).strip():
                        preview_rows.append(item)
        except Exception as e:
            messages.error(request, f'Failed to parse file for preview: {e}')
            preview_rows = []

        messages.success(request, f'File uploaded: {filename}. Parsed {len(preview_rows)} rows with last_name for preview.')
        audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) uploaded officer bulk event file {filename} parsed_rows={len(preview_rows)}")
        return render(request, 'officer/bulk_event_reg.html', {
            'preview_rows': preview_rows,
            'event': Event.objects.filter(status='active').order_by('-date').first()
        })

    return render(request, 'officer/bulk_event_reg.html')

def SAVE_BULK_EVENT_REG(request):
    """Persist uploaded preview rows into Bulk_Event_Reg.

    Expects POST with 'preview_json' containing the preview rows shown in the table.
    Only rows with non-empty last_name should be processed (already filtered client-side).
    """
    if request.method != 'POST':
        return redirect('bulk_event_reg_officer')

    preview_json = request.POST.get('preview_json')
    if not preview_json:
        messages.error(request, 'No preview data to save.')
        return redirect('bulk_event_reg_officer')

    try:
        # The template uses json_script; adapt to plain JSON if provided
        # If it contains a <script> tag, extract inner text; otherwise treat as JSON
        data_str = preview_json
        # Attempt to load JSON
        rows = json.loads(data_str)
        if not isinstance(rows, list):
            messages.error(request, 'Invalid preview data format.')
            return redirect('bulk_event_reg_officer')
    except Exception as e:
        messages.error(request, f'Failed to parse preview data: {e}')
        return redirect('bulk_event_reg_officer')

    # Fetch active event to associate registrations
    event = Event.objects.filter(status='active').order_by('-date').first()
    if not event:
        messages.error(request, 'No active event to associate registrations.')
        return redirect('bulk_event_reg_officer')

    saved = 0
    skipped = 0

    # Persist within a transaction for reliability
    try:
        with transaction.atomic():
            def as_bool(val):
                if val is None:
                    return False
                s = str(val).strip().lower()
                return s in ['yes','true','1','y','t']

            # Resolve Member.id for the current user (instead of CustomUser.id)
            member_id = (
                Member.objects.filter(admin_id=getattr(request.user, 'id', None))
                .values_list('id', flat=True)
                .first()
            )

            for r in rows:
                try:
                    # Only save fields that exist; map expected keys
                    reg = Bulk_Event_Reg.objects.create(
                        event_id=event.id,
                        # Store Member.id, since registered_by FK points to Member
                        registered_by_id=member_id,
                        last_name=r.get('last_name', '') or '',
                        first_name=r.get('first_name', '') or '',
                        middle=r.get('middle', '') or '',
                        contact_number=r.get('contact_number', '') or '',
                        email=r.get('email', '') or '',
                        attending_as=r.get('attending_as', '') or '',
                        is_competitor=as_bool(r.get('is_competitor', False)),
                        if_competitor=r.get('if_competitor', '') or '',
                        is_coach=as_bool(r.get('is_coach', False)),
                        if_coach=r.get('if_coach', '') or '',
                        tshirt_size=r.get('tshirt_size', '') or '',
                    )
                    # Attending records are synchronized automatically by the post_save
                    # handler on `Bulk_Event_Reg` in `app.models`, so no manual creation here.
                    saved += 1
                except Exception:
                    skipped += 1
                    continue
    except Exception as e:
        messages.error(request, f'Failed to save registrations: {e}')
        audit_logger.exception(f"Failed to save officer bulk registrations for user {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}): {e}")
        return redirect('bulk_event_reg_officer')

    # Do not modify event.available_slots when saving bulk registrations.
    messages.success(request, f'Saved {saved} registrations. Skipped {skipped}.')
    audit_logger.info(f"User {getattr(request.user, 'username', None)} (id={getattr(request.user, 'id', None)}) saved officer bulk registrations for event id={getattr(event, 'id', None)} saved={saved} skipped={skipped}")
    return redirect('bulk_event_reg_officer')


    
